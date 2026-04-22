import calendar
import re
from datetime import date, timedelta
from io import BytesIO

from django.db import transaction
from django.db.models import Prefetch, Q, Sum, Count, F
from django.db.models.functions import Coalesce
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils.timezone import localdate

from .models import Semester, Siswa, JenisPembayaran, Tagihan, Pembayaran, TransaksiPembayaran, KasKeluar
from .forms import (
    SiswaForm, JenisPembayaranForm, PembayaranMultiForm,
    SemesterForm, BulkTagihanForm, KasKeluarForm
)

MONTH_NAMES_ID = [
    'Januari',
    'Februari',
    'Maret',
    'April',
    'Mei',
    'Juni',
    'Juli',
    'Agustus',
    'September',
    'Oktober',
    'November',
    'Desember',
]


def get_active_semester():
    return Semester.objects.filter(aktif=True).first()


def get_current_semester(request, source='GET'):
    data = request.POST if source == 'POST' else request.GET
    semester_id = data.get('semester') or data.get('semester_id')
    if semester_id:
        semester = get_object_or_404(Semester, pk=semester_id)
        if not semester.aktif:
            semester.aktif = True
            semester.save()
        return semester
    return get_active_semester() or Semester.objects.first()


def semester_query_param(semester):
    return f"semester={semester.pk}" if semester else ""


def add_months(source_date, months):
    month_index = source_date.month - 1 + months
    target_year = source_date.year + month_index // 12
    target_month = month_index % 12 + 1
    target_day = min(source_date.day, calendar.monthrange(target_year, target_month)[1])
    return date(target_year, target_month, target_day)


def format_month_year_id(value):
    return f"{MONTH_NAMES_ID[value.month - 1]} {value.year}"


def get_semester_month_span(semester, fallback=None):
    if semester and semester.tanggal_mulai and semester.tanggal_selesai:
        return max(
            ((semester.tanggal_selesai.year - semester.tanggal_mulai.year) * 12)
            + (semester.tanggal_selesai.month - semester.tanggal_mulai.month)
            + 1,
            1,
        )
    return fallback or 1


def build_monthly_periods(semester, count, due_date=None):
    base_date = semester.tanggal_mulai if semester else localdate()
    due_base = due_date or None
    resolved_count = get_semester_month_span(semester, count)
    periods = []

    for index in range(resolved_count):
        period_date = add_months(base_date, index)
        periods.append({
            'urutan_periode': index + 1,
            'label': format_month_year_id(period_date),
            'jatuh_tempo_default': add_months(due_base, index) if due_base else None,
        })

    return periods


def build_tagihan_status(total_nominal, total_terbayar, has_tagihan):
    total_sisa = max(total_nominal - total_terbayar, 0)
    if not has_tagihan:
        return 'Belum Dibuat'
    if total_terbayar <= 0:
        return 'Belum Bayar'
    if total_sisa <= 0:
        return 'Lunas'
    return 'Cicilan'


def get_applicable_jenis_queryset_for_siswa(siswa, *, recurring_only=False, include_ids=None):
    include_ids = list(include_ids or [])
    grade = siswa.tingkat_kelas
    applicable_filter = Q(target_kelas='')
    if grade:
        applicable_filter |= Q(target_kelas=grade)

    active_filter = Q(aktif=True) & applicable_filter
    if recurring_only:
        active_filter &= Q(wajib_per_semester=True)

    queryset = JenisPembayaran.objects.filter(active_filter)
    if include_ids:
        queryset = JenisPembayaran.objects.filter(active_filter | Q(id__in=include_ids))

    return queryset.distinct().order_by('nama')


def create_auto_tagihan_for_siswa(siswa, semester=None):
    semester = semester or get_active_semester()
    result = {
        'semester': semester,
        'created_count': 0,
        'skipped_count': 0,
    }

    if not siswa or not siswa.aktif or not semester:
        return result

    jenis_list = get_applicable_jenis_queryset_for_siswa(siswa, recurring_only=True)

    with transaction.atomic():
        for jenis in jenis_list:
            if jenis.is_bulanan:
                for period in build_monthly_periods(semester, jenis.jumlah_bulan_per_semester):
                    _, created = Tagihan.objects.get_or_create(
                        siswa=siswa,
                        jenis=jenis,
                        semester=semester,
                        urutan_periode=period['urutan_periode'],
                        defaults={
                            'nominal': jenis.nominal_default,
                            'periode': period['label'],
                            'jatuh_tempo': period['jatuh_tempo_default'],
                        },
                    )
                    if created:
                        result['created_count'] += 1
                    else:
                        result['skipped_count'] += 1
            else:
                _, created = Tagihan.objects.get_or_create(
                    siswa=siswa,
                    jenis=jenis,
                    semester=semester,
                    urutan_periode=0,
                    defaults={
                        'nominal': jenis.nominal_default,
                    },
                )
                if created:
                    result['created_count'] += 1
                else:
                    result['skipped_count'] += 1

    return result


def can_delete_tagihan(tagihan):
    return not tagihan.pembayaran_set.exists()


def get_available_tagihan_for_payment(siswa=None, semester=None):
    queryset = Tagihan.objects.select_related('siswa', 'jenis', 'semester').order_by(
        'jenis__nama', 'urutan_periode', 'periode', 'pk'
    )
    if siswa:
        queryset = queryset.filter(siswa=siswa)
    if semester:
        queryset = queryset.filter(semester=semester)

    return [
        tagihan for tagihan in queryset
        if tagihan.sisa_tagihan > 0
    ]


def build_pembayaran_groups(tagihan_list, data=None, preselected_tagihan_id=None):
    groups = []
    grouped_by_jenis = {}

    for tagihan in tagihan_list:
        amount_value = ''
        is_selected = False

        if data is not None:
            amount_value = data.get(f'jumlah_tagihan_{tagihan.pk}', '').strip()
            is_selected = (
                data.get(f'pilih_tagihan_{tagihan.pk}') == '1' or
                bool(amount_value)
            )
        elif preselected_tagihan_id and str(tagihan.pk) == str(preselected_tagihan_id):
            amount_value = str(tagihan.sisa_tagihan)
            is_selected = True

        group = grouped_by_jenis.get(tagihan.jenis_id)
        if group is None:
            group = {
                'jenis': tagihan.jenis,
                'is_bulanan': tagihan.jenis.is_bulanan,
                'rows': [],
                'total_nominal': 0,
                'total_terbayar': 0,
                'total_sisa': 0,
            }
            grouped_by_jenis[tagihan.jenis_id] = group
            groups.append(group)

        row = {
            'tagihan': tagihan,
            'selected': is_selected,
            'jumlah_input': amount_value,
        }
        group['rows'].append(row)
        group['total_nominal'] += tagihan.nominal
        group['total_terbayar'] += tagihan.total_terbayar
        group['total_sisa'] += tagihan.sisa_tagihan

    for group in groups:
        group['jumlah_rows'] = len(group['rows'])
        group['periode_labels'] = [
            row['tagihan'].periode or row['tagihan'].semester.nama
            for row in group['rows']
        ]

    return groups


def build_tagihan_display_groups(tagihan_list):
    groups = []
    grouped = {}

    for tagihan in tagihan_list:
        pembayaran_items = list(
            tagihan.pembayaran_set.select_related('transaksi').all().order_by('-tanggal_bayar')
        ) if hasattr(tagihan, '_prefetched_objects_cache') else []
        is_monthly = tagihan.jenis.is_bulanan
        key = (
            tagihan.semester_id,
            tagihan.jenis_id,
        ) if is_monthly else (
            tagihan.semester_id,
            tagihan.jenis_id,
            tagihan.pk,
        )

        group = grouped.get(key)
        if group is None:
            group = {
                'jenis': tagihan.jenis,
                'semester': tagihan.semester,
                'is_bulanan': is_monthly,
                'rows': [],
                'total_nominal': 0,
                'total_terbayar': 0,
                'total_sisa': 0,
            }
            grouped[key] = group
            groups.append(group)

        group['rows'].append({
            'tagihan': tagihan,
            'pembayaran_items': pembayaran_items,
            'periode_label': tagihan.periode or tagihan.semester.nama,
        })
        group['total_nominal'] += tagihan.nominal
        group['total_terbayar'] += tagihan.total_terbayar
        group['total_sisa'] += tagihan.sisa_tagihan

    for group in groups:
        group['jumlah_rows'] = len(group['rows'])
        group['status_pembayaran'] = build_tagihan_status(
            group['total_nominal'],
            group['total_terbayar'],
            bool(group['rows']),
        )
        group['periode_summary'] = ', '.join(
            row['periode_label'] for row in group['rows']
        )

    return groups


def build_payment_receipt_groups(payment_items, seluruh_tagihan=None):
    groups = []
    grouped = {}
    paid_tagihan_ids_by_group = {}

    for payment in payment_items:
        tagihan = payment.tagihan
        is_monthly = tagihan.jenis.is_bulanan
        key = (
            tagihan.semester_id,
            tagihan.jenis_id,
        ) if is_monthly else (
            tagihan.semester_id,
            tagihan.jenis_id,
            tagihan.pk,
        )

        group = grouped.get(key)
        if group is None:
            group = {
                'jenis': tagihan.jenis,
                'semester': tagihan.semester,
                'is_bulanan': is_monthly,
                'rows': [],
                'remaining_rows': [],
                'total_bayar': 0,
                'total_sisa_setelah_bayar': 0,
                'total_sisa_item': 0,
            }
            grouped[key] = group
            groups.append(group)
            paid_tagihan_ids_by_group[key] = set()

        group['rows'].append({
            'payment': payment,
            'tagihan': tagihan,
            'periode_label': tagihan.periode or tagihan.semester.nama,
            'sisa_setelah_bayar': tagihan.sisa_tagihan,
        })
        group['total_bayar'] += payment.jumlah_bayar
        group['total_sisa_setelah_bayar'] += tagihan.sisa_tagihan
        paid_tagihan_ids_by_group[key].add(tagihan.id)

    if seluruh_tagihan is not None:
        for tagihan in seluruh_tagihan:
            is_monthly = tagihan.jenis.is_bulanan
            key = (
                tagihan.semester_id,
                tagihan.jenis_id,
            ) if is_monthly else (
                tagihan.semester_id,
                tagihan.jenis_id,
                tagihan.pk,
            )

            group = grouped.get(key)
            if group is None:
                continue

            group['total_sisa_item'] += tagihan.sisa_tagihan
            if tagihan.id in paid_tagihan_ids_by_group[key] or tagihan.sisa_tagihan <= 0:
                continue

            group['remaining_rows'].append({
                'tagihan': tagihan,
                'periode_label': tagihan.periode or tagihan.semester.nama,
                'sisa_tagihan': tagihan.sisa_tagihan,
            })

    for group in groups:
        group['jumlah_rows'] = len(group['rows'])
        group['periode_summary'] = ', '.join(
            row['periode_label'] for row in group['rows']
        )
        group['remaining_count'] = len(group['remaining_rows'])

    return groups


def extract_tingkat_from_kelas_label(kelas_label):
    normalized = re.sub(r'\s+', '', (kelas_label or '').strip().upper())
    if not normalized:
        return ''

    roman_map = (
        ('VIII', '8'),
        ('VII', '7'),
        ('IX', '9'),
    )
    for prefix, grade in roman_map:
        if normalized.startswith(prefix):
            return grade

    match = re.match(r'(\d+)', normalized)
    return match.group(1) if match else ''


def get_applicable_jenis_queryset_for_kelas(kelas_label):
    grade = extract_tingkat_from_kelas_label(kelas_label)
    queryset = JenisPembayaran.objects.filter(aktif=True)
    if grade:
        queryset = queryset.filter(Q(target_kelas='') | Q(target_kelas=grade))
    else:
        queryset = queryset.filter(target_kelas='')
    return queryset.order_by('nama')


def build_tagihan_group_summaries(tagihan_items):
    groups = []
    grouped = {}

    for tagihan in tagihan_items:
        group = grouped.get(tagihan.jenis_id)
        if group is None:
            group = {
                'jenis': tagihan.jenis,
                'rows': [],
                'total_nominal': 0,
                'total_terbayar': 0,
                'total_sisa': 0,
                'periode_labels': [],
                'jumlah_lunas': 0,
            }
            grouped[tagihan.jenis_id] = group
            groups.append(group)

        group['rows'].append(tagihan)
        group['total_nominal'] += tagihan.nominal
        group['total_terbayar'] += tagihan.total_terbayar
        group['total_sisa'] += tagihan.sisa_tagihan
        group['periode_labels'].append(tagihan.periode or tagihan.semester.nama)
        if tagihan.sisa_tagihan <= 0:
            group['jumlah_lunas'] += 1

    for group in groups:
        group['jumlah_item'] = len(group['rows'])
        group['periode_summary'] = ', '.join(group['periode_labels'])
        if group['total_sisa'] <= 0:
            group['status'] = 'Lunas'
            group['status_tone'] = 'green'
        elif group['total_terbayar'] > 0:
            group['status'] = 'Sudah Bayar'
            group['status_tone'] = 'amber'
        else:
            group['status'] = 'Belum Bayar'
            group['status_tone'] = 'red'

        if group['jenis'].is_bulanan:
            group['detail_status'] = (
                f"{group['jumlah_lunas']}/{group['jumlah_item']} bulan lunas"
            )
        else:
            group['detail_status'] = (
                'Sudah lunas'
                if group['status'] == 'Lunas'
                else 'Sudah ada pembayaran' if group['status'] == 'Sudah Bayar'
                else 'Belum ada pembayaran'
            )

    return groups


def format_tagihan_group_details(tagihan_items):
    groups = build_tagihan_group_summaries(tagihan_items)
    details = []
    for group in groups:
        if group['jenis'].is_bulanan:
            details.append(
                f"{group['jenis'].nama} [{group['periode_summary']}]: {group['status']} "
                f"(sisa {format_rupiah(group['total_sisa'])})"
            )
        else:
            periode_label = group['periode_summary'] or '-'
            details.append(
                f"{group['jenis'].nama} ({periode_label}): {group['status']} "
                f"(sisa {format_rupiah(group['total_sisa'])})"
            )
    return '; '.join(details)


def format_outstanding_group_details(tagihan_items):
    groups = build_tagihan_group_summaries(tagihan_items)
    details = []
    for group in groups:
        if group['total_sisa'] <= 0:
            continue
        if group['jenis'].is_bulanan:
            details.append(
                f"{group['jenis'].nama} [{group['periode_summary']}]: {format_rupiah(group['total_sisa'])}"
            )
        else:
            periode_label = group['periode_summary'] or '-'
            details.append(
                f"{group['jenis'].nama} ({periode_label}): {format_rupiah(group['total_sisa'])}"
            )
    return '; '.join(details)


def build_payment_report_rows(payment_items):
    rows = []
    grouped = {}

    for item in payment_items:
        transaction_key = (
            f"trx-{item.transaksi_id}"
            if item.transaksi_id and item.transaksi
            else f"single-{item.pk}"
        )
        key = (
            transaction_key,
            item.tagihan.siswa_id,
            item.tagihan.jenis_id,
        ) if item.tagihan.jenis.is_bulanan else (
            transaction_key,
            item.pk,
        )

        row = grouped.get(key)
        if row is None:
            row = {
                'tanggal_bayar': item.tanggal_bayar,
                'kode_transaksi': item.transaksi.kode_transaksi if item.transaksi_id and item.transaksi else f"PBY-{item.pk:05d}",
                'siswa': item.tagihan.siswa,
                'jenis': item.tagihan.jenis,
                'metode': item.metode or '-',
                'jumlah_bayar': 0,
                'periode_labels': [],
                'item_count': 0,
            }
            grouped[key] = row
            rows.append(row)

        row['jumlah_bayar'] += item.jumlah_bayar
        row['periode_labels'].append(item.tagihan.periode or item.tagihan.semester.nama)
        row['item_count'] += 1

    for row in rows:
        row['periode_summary'] = ', '.join(row['periode_labels'])

    return rows


def format_rupiah(value):
    return f"Rp {int(value or 0):,}"


def increment_tahun_ajaran(tahun_ajaran):
    parts = [item.strip() for item in (tahun_ajaran or '').split('/')]
    if len(parts) == 2 and all(part.isdigit() for part in parts):
        return f"{int(parts[0]) + 1}/{int(parts[1]) + 1}"

    current_year = localdate().year
    return f"{current_year}/{current_year + 1}"


def build_semester_suggestion():
    latest_semester = Semester.objects.order_by('-tanggal_selesai', '-id').first()
    today = localdate()

    if latest_semester:
        next_semester_name = 'Genap' if latest_semester.semester.lower() == 'ganjil' else 'Ganjil'
        next_tahun_ajaran = (
            latest_semester.tahun_ajaran
            if next_semester_name == 'Genap'
            else increment_tahun_ajaran(latest_semester.tahun_ajaran)
        )
        duration_days = max((latest_semester.tanggal_selesai - latest_semester.tanggal_mulai).days, 180)
        tanggal_mulai = latest_semester.tanggal_selesai + timedelta(days=1)
        tanggal_selesai = tanggal_mulai + timedelta(days=duration_days)
        source_label = f"Disarankan dari semester terakhir: {latest_semester.nama}"
    else:
        if today.month >= 7:
            next_semester_name = 'Ganjil'
            next_tahun_ajaran = f"{today.year}/{today.year + 1}"
            tanggal_mulai = date(today.year, 7, 1)
            tanggal_selesai = date(today.year, 12, 31)
        else:
            next_semester_name = 'Genap'
            next_tahun_ajaran = f"{today.year - 1}/{today.year}"
            tanggal_mulai = date(today.year, 1, 1)
            tanggal_selesai = date(today.year, 6, 30)
        source_label = "Disarankan dari pola kalender semester madrasah"

    preview_count = get_semester_month_span(
        type('SemesterPreview', (), {
            'tanggal_mulai': tanggal_mulai,
            'tanggal_selesai': tanggal_selesai,
        })(),
        6,
    )
    month_preview = [
        format_month_year_id(add_months(tanggal_mulai, index))
        for index in range(preview_count)
    ]

    return {
        'source_label': source_label,
        'nama': f"{next_semester_name} {next_tahun_ajaran}",
        'tahun_ajaran': next_tahun_ajaran,
        'semester': next_semester_name,
        'tanggal_mulai': tanggal_mulai,
        'tanggal_selesai': tanggal_selesai,
        'month_preview': month_preview,
        'month_count': preview_count,
    }


def build_dashboard_notifications(*, today, semester_aktif, dashboard_tagihan, total_tunggakan, jumlah_siswa_belum_punya_tagihan):
    notifications = []

    if not semester_aktif:
        notifications.append({
            'tone': 'danger',
            'title': 'Semester aktif belum ditentukan',
            'description': 'Segera atur semester aktif agar tagihan, pembayaran, dan laporan mengikuti periode madrasah yang benar.',
            'action_label': 'Atur Semester',
            'action_url': '/bendahara/semester/',
        })
        return notifications

    if today < semester_aktif.tanggal_mulai:
        notifications.append({
            'tone': 'warning',
            'title': 'Semester aktif belum dimulai',
            'description': f"Semester {semester_aktif.nama} baru dimulai pada {semester_aktif.tanggal_mulai.strftime('%d-%m-%Y')}. Cek kembali kesiapan tagihan awal periode.",
            'action_label': 'Cek Tagihan',
            'action_url': f"/bendahara/tagihan/?semester={semester_aktif.id}",
        })
    elif today > semester_aktif.tanggal_selesai:
        notifications.append({
            'tone': 'danger',
            'title': 'Semester aktif sudah melewati akhir periode',
            'description': f"Semester {semester_aktif.nama} berakhir pada {semester_aktif.tanggal_selesai.strftime('%d-%m-%Y')}. Sebaiknya siapkan semester baru sesuai kaldik madrasah.",
            'action_label': 'Buat Semester Baru',
            'action_url': '/bendahara/semester/tambah/',
        })
    else:
        sisa_hari_semester = (semester_aktif.tanggal_selesai - today).days
        if sisa_hari_semester <= 21:
            notifications.append({
                'tone': 'warning',
                'title': 'Semester aktif segera berakhir',
                'description': f"Tersisa {sisa_hari_semester} hari menuju akhir semester {semester_aktif.nama}. Mulai siapkan periode berikutnya dan rapikan tunggakan.",
                'action_label': 'Lihat Laporan Semester',
                'action_url': f"/bendahara/laporan/semester/?semester={semester_aktif.id}",
            })

    overdue_tagihan = [
        tagihan for tagihan in dashboard_tagihan
        if tagihan.sisa_tagihan > 0 and tagihan.jatuh_tempo and tagihan.jatuh_tempo < today
    ]
    if overdue_tagihan:
        notifications.append({
            'tone': 'danger',
            'title': f"{len(overdue_tagihan)} tagihan melewati jatuh tempo",
            'description': f"Masih ada tagihan lewat jatuh tempo dengan total sisa {format_rupiah(sum(item.sisa_tagihan for item in overdue_tagihan))}.",
            'action_label': 'Lihat Pembayaran',
            'action_url': f"/bendahara/pembayaran/?semester={semester_aktif.id}",
        })

    if jumlah_siswa_belum_punya_tagihan > 0:
        notifications.append({
            'tone': 'info',
            'title': f"{jumlah_siswa_belum_punya_tagihan} siswa belum punya tagihan",
            'description': 'Data siswa aktif sudah ada, tetapi sebagian belum memiliki tagihan pada semester berjalan.',
            'action_label': 'Kelola Tagihan',
            'action_url': f"/bendahara/tagihan/?semester={semester_aktif.id}",
        })

    if total_tunggakan > 0:
        notifications.append({
            'tone': 'warning',
            'title': 'Tunggakan semester masih perlu dipantau',
            'description': f"Total sisa tagihan pada semester aktif saat ini adalah {format_rupiah(total_tunggakan)}.",
            'action_label': 'Buka Laporan Tunggakan',
            'action_url': f"/bendahara/laporan/tunggakan/?semester={semester_aktif.id}",
        })

    if not notifications:
        notifications.append({
            'tone': 'success',
            'title': 'Kondisi semester aktif cukup terkendali',
            'description': 'Tidak ada peringatan utama saat ini. Lanjutkan monitoring pembayaran dan kas sekolah seperti biasa.',
            'action_label': 'Lihat Monitoring',
            'action_url': '/monitoring-bendahara/',
        })

    return notifications[:4]


def build_report_filename(prefix, extension):
    return f"{prefix}_{localdate().strftime('%Y%m%d')}.{extension}"


def get_report_export_format(request):
    export_format = (request.GET.get('export') or '').strip().lower()
    return export_format if export_format in {'pdf', 'xlsx'} else ''


def get_kas_summary(semester=None):
    pembayaran_queryset = Pembayaran.objects.select_related('tagihan__semester')
    kas_keluar_queryset = KasKeluar.objects.select_related('semester')

    if semester:
        pembayaran_queryset = pembayaran_queryset.filter(tagihan__semester=semester)
        kas_keluar_queryset = kas_keluar_queryset.filter(semester=semester)

    total_kas_masuk = sum(item.jumlah_bayar for item in pembayaran_queryset)
    total_kas_keluar = sum(item.jumlah for item in kas_keluar_queryset)
    return {
        'total_kas_masuk': total_kas_masuk,
        'total_kas_keluar': total_kas_keluar,
        'saldo_kas': total_kas_masuk - total_kas_keluar,
    }


def get_cash_position_summary(semester=None):
    tagihan_queryset = Tagihan.objects.select_related('semester')
    pembayaran_queryset = Pembayaran.objects.select_related('tagihan__semester')
    kas_keluar_queryset = KasKeluar.objects.select_related('semester')

    if semester:
        tagihan_queryset = tagihan_queryset.filter(semester=semester)
        pembayaran_queryset = pembayaran_queryset.filter(tagihan__semester=semester)
        kas_keluar_queryset = kas_keluar_queryset.filter(semester=semester)

    total_target_masuk = sum(tagihan.nominal for tagihan in tagihan_queryset)
    total_realisasi_masuk = sum(item.jumlah_bayar for item in pembayaran_queryset)
    total_pengeluaran = sum(item.jumlah for item in kas_keluar_queryset)

    return {
        'total_target_masuk': total_target_masuk,
        'total_realisasi_masuk': total_realisasi_masuk,
        'total_pengeluaran': total_pengeluaran,
        'saldo_aktual': total_realisasi_masuk - total_pengeluaran,
        'sisa_target_belum_masuk': max(total_target_masuk - total_realisasi_masuk, 0),
    }


def export_report_excel(*, title, filename, sheet_name, filter_rows, summary_rows, headers, data_rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31]

    last_column = max(len(headers), 1)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    worksheet['A1'] = title
    worksheet['A1'].font = Font(size=15, bold=True)
    worksheet['A1'].alignment = Alignment(horizontal='center')

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    worksheet['A2'] = 'MTs Sunan Kalijaga'
    worksheet['A2'].font = Font(size=11, bold=True)
    worksheet['A2'].alignment = Alignment(horizontal='center')

    worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_column)
    worksheet['A3'] = f"Tanggal cetak: {localdate().strftime('%d-%m-%Y')}"
    worksheet['A3'].alignment = Alignment(horizontal='center')

    current_row = 5
    if filter_rows:
        worksheet.cell(row=current_row, column=1, value='Filter Laporan')
        worksheet.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        for label, value in filter_rows:
            worksheet.cell(row=current_row, column=1, value=label)
            worksheet.cell(row=current_row, column=2, value=value)
            current_row += 1
        current_row += 1

    if summary_rows:
        worksheet.cell(row=current_row, column=1, value='Ringkasan')
        worksheet.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        for label, value in summary_rows:
            worksheet.cell(row=current_row, column=1, value=label)
            worksheet.cell(row=current_row, column=2, value=value)
            current_row += 1
        current_row += 1

    header_fill = PatternFill(fill_type='solid', fgColor='1D4ED8')
    header_font = Font(color='FFFFFF', bold=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=current_row, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    data_start_row = current_row + 1
    for row_index, row in enumerate(data_rows, start=data_start_row):
        for column_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
        if row_index % 2 == 0:
            for column_index in range(1, len(headers) + 1):
                worksheet.cell(row=row_index, column=column_index).fill = PatternFill(
                    fill_type='solid',
                    fgColor='F8FAFC',
                )

    worksheet.freeze_panes = f"A{data_start_row}"

    for column_index in range(1, len(headers) + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0
        for cell in worksheet[column_letter]:
            cell_value = '' if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_report_pdf(*, title, filename, filter_rows, summary_rows, headers, data_rows):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = BytesIO()
    page_size = landscape(A4) if len(headers) > 6 else A4
    document = SimpleDocTemplate(
        output,
        pagesize=page_size,
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"<b>{title}</b>", styles['Title']),
        Paragraph("MTs Sunan Kalijaga", styles['Heading3']),
        Paragraph(f"Tanggal cetak: {localdate().strftime('%d-%m-%Y')}", styles['BodyText']),
        Spacer(1, 12),
    ]

    if filter_rows:
        elements.append(Paragraph("<b>Filter Laporan</b>", styles['Heading4']))
        filter_table = Table([[label, value] for label, value in filter_rows], hAlign='LEFT', colWidths=[140, 320])
        filter_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.extend([filter_table, Spacer(1, 12)])

    if summary_rows:
        elements.append(Paragraph("<b>Ringkasan</b>", styles['Heading4']))
        summary_table = Table([[label, value] for label, value in summary_rows], hAlign='LEFT', colWidths=[160, 180])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#EFF6FF')),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#93C5FD')),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#BFDBFE')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.extend([summary_table, Spacer(1, 12)])

    table_data = [headers] + data_rows
    col_width = (document.width / max(len(headers), 1))
    report_table = Table(table_data, repeatRows=1, colWidths=[col_width] * len(headers))
    report_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1D4ED8')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CBD5E1')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(report_table)

    document.build(elements)
    output.seek(0)

    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_report_response(*, request, title, filename_prefix, sheet_name, filter_rows, summary_rows, headers, data_rows):
    export_format = get_report_export_format(request)
    if not export_format:
        return None

    if export_format == 'xlsx':
        return export_report_excel(
            title=title,
            filename=build_report_filename(filename_prefix, 'xlsx'),
            sheet_name=sheet_name,
            filter_rows=filter_rows,
            summary_rows=summary_rows,
            headers=headers,
            data_rows=data_rows,
        )

    return export_report_pdf(
        title=title,
        filename=build_report_filename(filename_prefix, 'pdf'),
        filter_rows=filter_rows,
        summary_rows=summary_rows,
        headers=headers,
        data_rows=data_rows,
    )


def render_pembayaran_receipt_response(*, transaksi=None, pembayaran=None):
    if transaksi is not None:
        payment_items = list(
            transaksi.pembayaran_set.select_related(
                'tagihan__jenis',
                'tagihan__semester',
                'tagihan__siswa',
            ).order_by('tagihan__jenis__nama', 'tagihan__urutan_periode', 'pk')
        )
        siswa = transaksi.siswa
        nomor_transaksi = transaksi.kode_transaksi or f"TRX-{transaksi.pk}"
        tanggal_bayar = transaksi.tanggal_bayar
        metode = transaksi.metode
        keterangan = transaksi.keterangan
        total_bayar = sum(item.jumlah_bayar for item in payment_items)
        semester = transaksi.semester
        filename = f'struk_{nomor_transaksi}.html'
    elif pembayaran is not None:
        payment_items = [pembayaran]
        siswa = pembayaran.tagihan.siswa
        nomor_transaksi = f"PBY-{pembayaran.pk:05d}"
        tanggal_bayar = pembayaran.tanggal_bayar
        metode = pembayaran.metode
        keterangan = pembayaran.keterangan
        total_bayar = pembayaran.jumlah_bayar
        semester = pembayaran.tagihan.semester
        filename = f'kwitansi_pembayaran_{pembayaran.pk}.html'
    else:
        raise ValueError("transaksi atau pembayaran wajib disediakan")

    seluruh_tagihan_queryset = Tagihan.objects.filter(siswa=siswa)
    if semester:
        seluruh_tagihan_queryset = seluruh_tagihan_queryset.filter(semester=semester)

    seluruh_tagihan = list(
        seluruh_tagihan_queryset.select_related('jenis', 'semester').prefetch_related('pembayaran_set')
    )
    total_sisa_setelah_bayar = sum(tagihan.sisa_tagihan for tagihan in seluruh_tagihan)
    status_pelunasan = 'Lunas' if total_sisa_setelah_bayar <= 0 else 'Kurang Bayar'
    payment_groups = build_payment_receipt_groups(payment_items, seluruh_tagihan=seluruh_tagihan)

    html = render_to_string('bendahara/pembayaran_receipt.html', {
        'transaksi': transaksi,
        'pembayaran': pembayaran,
        'payment_items': payment_items,
        'payment_groups': payment_groups,
        'siswa': siswa,
        'nomor_transaksi': nomor_transaksi,
        'tanggal_bayar': tanggal_bayar,
        'metode': metode,
        'keterangan': keterangan,
        'total_bayar': total_bayar,
        'total_sisa_setelah_bayar': total_sisa_setelah_bayar,
        'status_pelunasan': status_pelunasan,
        'semester': semester,
    })

    response = HttpResponse(html, content_type='text/html; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ================= LOGIN =================
def login_bendahara(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)

        if user:
            login(request, user)
            messages.success(request, f"Selamat datang, {user.username}. Berikut ringkasan penting operasional bendahara hari ini.")
            return redirect('bendahara:dashboard')

        messages.error(request, "Username atau password belum sesuai. Silakan periksa kembali.")

    return render(request, 'bendahara/login.html')


def logout_bendahara(request):
    logout(request)
    return redirect('bendahara:login')


# ================= DASHBOARD =================
def dashboard(request):
    today = localdate()

    semester_aktif = get_active_semester()

    if semester_aktif:
        tagihan_queryset = Tagihan.objects.filter(semester=semester_aktif)
        pembayaran_queryset = Pembayaran.objects.filter(tagihan__semester=semester_aktif)
    else:
        tagihan_queryset = Tagihan.objects.none()
        pembayaran_queryset = Pembayaran.objects.none()
    
    siswa_objects = (
        Siswa.objects.prefetch_related(
            Prefetch(
                'tagihan_set',
                queryset=tagihan_queryset.select_related('jenis', 'semester').prefetch_related('pembayaran_set'),
            )
        )
        .order_by('nama')
    )

    total_siswa_aktif = Siswa.objects.filter(aktif=True).count()
    total_jenis_pembayaran_aktif = JenisPembayaran.objects.filter(aktif=True).count()
    total_tagihan = tagihan_queryset.count()
    total_pembayaran = pembayaran_queryset.count()
    pembayaran_hari_ini = pembayaran_queryset.filter(tanggal_bayar__date=today)
    total_pembayaran_hari_ini = sum(item.jumlah_bayar for item in pembayaran_hari_ini)

    total_tagihan_nominal = 0
    total_terbayar = 0
    siswa_tunggakan = []
    siswa_belum_punya_tagihan = []

    for siswa in siswa_objects:
        tagihan_items = list(siswa.tagihan_set.all())
        siswa_total_nominal = sum(tagihan.nominal for tagihan in tagihan_items)
        siswa_total_terbayar = sum(tagihan.total_terbayar for tagihan in tagihan_items)
        siswa_total_sisa = max(siswa_total_nominal - siswa_total_terbayar, 0)

        total_tagihan_nominal += siswa_total_nominal
        total_terbayar += siswa_total_terbayar

        if not tagihan_items:
            siswa_belum_punya_tagihan.append(siswa)

        if siswa_total_sisa > 0:
            siswa_tunggakan.append({
                'id': siswa.id,
                'nis': siswa.nis,
                'nama': siswa.nama,
                'kelas': siswa.kelas,
                'jumlah_tagihan': len(tagihan_items),
                'total_sisa': siswa_total_sisa,
            })

    siswa_tunggakan.sort(key=lambda item: item['total_sisa'], reverse=True)

    recent_pembayaran = (
        pembayaran_queryset.select_related('tagihan__siswa', 'tagihan__jenis', 'tagihan__semester')
        .order_by('-tanggal_bayar')[:5]
    )

    dashboard_tagihan = list(
        tagihan_queryset.select_related('siswa', 'jenis', 'semester')
        .prefetch_related('pembayaran_set')
        .order_by('siswa__nama', 'jenis__nama', 'urutan_periode')
    )
    tagihan_kurang_bayar = [
        tagihan for tagihan in dashboard_tagihan
        if tagihan.sisa_tagihan > 0
    ]
    tagihan_sudah_dibayar = [
        tagihan for tagihan in dashboard_tagihan
        if tagihan.total_terbayar > 0
    ]
    tagihan_kurang_bayar.sort(key=lambda tagihan: tagihan.sisa_tagihan, reverse=True)
    tagihan_sudah_dibayar.sort(key=lambda tagihan: tagihan.total_terbayar, reverse=True)

    jenis_dashboard_cards = []
    active_jenis_list = list(JenisPembayaran.objects.filter(aktif=True).order_by('nama'))
    tagihan_by_jenis = {}
    for tagihan in dashboard_tagihan:
        tagihan_by_jenis.setdefault(tagihan.jenis_id, []).append(tagihan)

    for jenis in active_jenis_list:
        jenis_tagihan = tagihan_by_jenis.get(jenis.id, [])
        total_nominal_jenis = sum(tagihan.nominal for tagihan in jenis_tagihan)
        total_terbayar_jenis = sum(tagihan.total_terbayar for tagihan in jenis_tagihan)
        total_sisa_jenis = sum(tagihan.sisa_tagihan for tagihan in jenis_tagihan)
        jenis_dashboard_cards.append({
            'jenis': jenis,
            'jumlah_item': len(jenis_tagihan),
            'total_nominal': total_nominal_jenis,
            'total_terbayar': total_terbayar_jenis,
            'total_sisa': total_sisa_jenis,
        })

    kas_summary = get_kas_summary(semester_aktif)

    context = {
        'today': today,
        'semester_aktif': semester_aktif,
        'total_siswa_aktif': total_siswa_aktif,
        'total_jenis_pembayaran_aktif': total_jenis_pembayaran_aktif,
        'total_tagihan': total_tagihan,
        'total_pembayaran': total_pembayaran,
        'total_tagihan_nominal': total_tagihan_nominal,
        'total_terbayar': total_terbayar,
        'total_tunggakan': max(total_tagihan_nominal - total_terbayar, 0),
        'total_pembayaran_hari_ini': total_pembayaran_hari_ini,
        'jumlah_transaksi_hari_ini': pembayaran_hari_ini.count(),
        'jumlah_siswa_nunggak': len(siswa_tunggakan),
        'jumlah_siswa_belum_punya_tagihan': len(siswa_belum_punya_tagihan),
        'siswa_tunggakan': siswa_tunggakan[:5],
        'siswa_belum_punya_tagihan': siswa_belum_punya_tagihan[:5],
        'recent_pembayaran': recent_pembayaran,
        'tagihan_kurang_bayar': tagihan_kurang_bayar[:8],
        'tagihan_sudah_dibayar': tagihan_sudah_dibayar[:8],
        'jumlah_tagihan_kurang_bayar': len(tagihan_kurang_bayar),
        'jumlah_tagihan_sudah_dibayar': len(tagihan_sudah_dibayar),
        'jenis_dashboard_cards': jenis_dashboard_cards[:6],
        'total_kas_masuk_semester': kas_summary['total_kas_masuk'],
        'total_kas_keluar_semester': kas_summary['total_kas_keluar'],
        'saldo_kas_semester': kas_summary['saldo_kas'],
        'dashboard_notifications': build_dashboard_notifications(
            today=today,
            semester_aktif=semester_aktif,
            dashboard_tagihan=dashboard_tagihan,
            total_tunggakan=max(total_tagihan_nominal - total_terbayar, 0),
            jumlah_siswa_belum_punya_tagihan=len(siswa_belum_punya_tagihan),
        ),
    }
    return render(request, 'bendahara/dashboard.html', context)


# ================= KAS SEKOLAH =================
def kas_sekolah(request):
    current_semester = get_current_semester(request, source='POST' if request.method == 'POST' else 'GET')
    selected_semester = current_semester

    if request.method == 'POST':
        form = KasKeluarForm(request.POST)
        if form.is_valid():
            kas_keluar = form.save(commit=False)
            if not kas_keluar.semester and current_semester:
                kas_keluar.semester = current_semester
            kas_keluar.save()
            messages.success(request, f"Pengeluaran {kas_keluar.kode_pengeluaran} berhasil dicatat.")
            url = redirect('bendahara:kas_sekolah')
            if kas_keluar.semester:
                url['Location'] += f'?{semester_query_param(kas_keluar.semester)}'
            return url
        messages.error(request, "Pengeluaran belum tersimpan. Periksa kembali data Anda.")
    else:
        initial = {'semester': current_semester.pk} if current_semester else None
        form = KasKeluarForm(initial=initial)

    semester_list = Semester.objects.all()

    pembayaran_queryset = Pembayaran.objects.select_related(
        'tagihan__siswa',
        'tagihan__jenis',
        'tagihan__semester',
        'transaksi',
    ).order_by('-tanggal_bayar', '-pk')
    kas_keluar_queryset = KasKeluar.objects.select_related('semester', 'jenis_pembayaran').order_by('-tanggal_pengeluaran', '-id')

    if selected_semester:
        pembayaran_queryset = pembayaran_queryset.filter(tagihan__semester=selected_semester)
        kas_keluar_queryset = kas_keluar_queryset.filter(semester=selected_semester)

    pembayaran_items = list(pembayaran_queryset)
    kas_keluar_items = list(kas_keluar_queryset)
    cash_position_filtered = get_cash_position_summary(selected_semester)
    cash_position_overall = get_cash_position_summary()

    kategori_summary_map = {}
    for item in kas_keluar_items:
        kategori_summary_map[item.kategori] = kategori_summary_map.get(item.kategori, 0) + item.jumlah
    kategori_summary = [
        {'kategori': kategori, 'total': total}
        for kategori, total in sorted(kategori_summary_map.items(), key=lambda value: value[1], reverse=True)
    ]

    jenis_summary_map = {}
    jenis_queryset = JenisPembayaran.objects.filter(aktif=True).order_by('nama')
    for jenis in jenis_queryset:
        jenis_summary_map[jenis.id] = {
            'jenis': jenis,
            'target': 0,
            'masuk': 0,
            'keluar': 0,
            'sisa_tagihan': 0,
            'saldo_dana': 0,
        }

    tagihan_for_summary = Tagihan.objects.select_related('jenis', 'semester').prefetch_related('pembayaran_set')
    if selected_semester:
        tagihan_for_summary = tagihan_for_summary.filter(semester=selected_semester)

    for tagihan in tagihan_for_summary:
        if tagihan.jenis_id not in jenis_summary_map:
            jenis_summary_map[tagihan.jenis_id] = {
                'jenis': tagihan.jenis,
                'target': 0,
                'masuk': 0,
                'keluar': 0,
                'sisa_tagihan': 0,
                'saldo_dana': 0,
            }
        jenis_summary_map[tagihan.jenis_id]['target'] += tagihan.nominal
        jenis_summary_map[tagihan.jenis_id]['masuk'] += tagihan.total_terbayar
        jenis_summary_map[tagihan.jenis_id]['sisa_tagihan'] += tagihan.sisa_tagihan

    for item in kas_keluar_items:
        if not item.jenis_pembayaran_id:
            continue
        if item.jenis_pembayaran_id not in jenis_summary_map:
            jenis_summary_map[item.jenis_pembayaran_id] = {
                'jenis': item.jenis_pembayaran,
                'target': 0,
                'masuk': 0,
                'keluar': 0,
                'sisa_tagihan': 0,
                'saldo_dana': 0,
            }
        jenis_summary_map[item.jenis_pembayaran_id]['keluar'] += item.jumlah

    jenis_keuangan_rows = []
    for item in jenis_summary_map.values():
        if item['target'] <= 0 and item['masuk'] <= 0 and item['keluar'] <= 0:
            continue
        item['saldo_dana'] = item['masuk'] - item['keluar']
        jenis_keuangan_rows.append(item)
    jenis_keuangan_rows.sort(key=lambda item: item['target'], reverse=True)

    total_kas_masuk_hari_ini = sum(
        item.jumlah_bayar for item in pembayaran_items if item.tanggal_bayar.date() == localdate()
    )
    total_kas_keluar_hari_ini = sum(
        item.jumlah for item in kas_keluar_items if item.tanggal_pengeluaran == localdate()
    )

    context = {
        'form': form,
        'semester_list': semester_list,
        'current_semester': selected_semester,
        'total_target_masuk': cash_position_filtered['total_target_masuk'],
        'total_realisasi_masuk': cash_position_filtered['total_realisasi_masuk'],
        'total_pengeluaran': cash_position_filtered['total_pengeluaran'],
        'saldo_aktual': cash_position_filtered['saldo_aktual'],
        'sisa_target_belum_masuk': cash_position_filtered['sisa_target_belum_masuk'],
        'total_target_masuk_keseluruhan': cash_position_overall['total_target_masuk'],
        'total_realisasi_masuk_keseluruhan': cash_position_overall['total_realisasi_masuk'],
        'total_pengeluaran_keseluruhan': cash_position_overall['total_pengeluaran'],
        'saldo_aktual_keseluruhan': cash_position_overall['saldo_aktual'],
        'sisa_target_belum_masuk_keseluruhan': cash_position_overall['sisa_target_belum_masuk'],
        'total_kas_masuk_hari_ini': total_kas_masuk_hari_ini,
        'total_kas_keluar_hari_ini': total_kas_keluar_hari_ini,
        'jumlah_transaksi_masuk': len(pembayaran_items),
        'jumlah_pengeluaran': len(kas_keluar_items),
        'pembayaran_items': pembayaran_items,
        'kas_keluar_items': kas_keluar_items,
        'kategori_summary': kategori_summary[:6],
        'jenis_keuangan_rows': jenis_keuangan_rows[:8],
    }
    return render(request, 'bendahara/kas_sekolah.html', context)


# ================= SEMESTER =================
def semester_list(request):
    semester_list = Semester.objects.all()
    return render(request, 'bendahara/semester_list.html', {
        'semester_list': semester_list
    })


def semester_create(request):
    semester_suggestion = build_semester_suggestion()

    if request.method == 'POST':
        form = SemesterForm(request.POST)
        if form.is_valid():
            semester = form.save()
            messages.success(request, "Semester berhasil ditambahkan!")
            if semester.aktif:
                messages.info(request, f"{semester.nama} sekarang menjadi semester aktif.")
            return redirect('bendahara:semester_list')
        else:
            messages.error(request, "Form tidak valid, cek kembali input!")
    else:
        form = SemesterForm(initial={
            'nama': semester_suggestion['nama'],
            'tahun_ajaran': semester_suggestion['tahun_ajaran'],
            'semester': semester_suggestion['semester'],
            'tanggal_mulai': semester_suggestion['tanggal_mulai'],
            'tanggal_selesai': semester_suggestion['tanggal_selesai'],
            'aktif': True,
        })
    
    return render(request, 'bendahara/semester_form.html', {
        'form': form,
        'mode': 'create',
        'semester_suggestion': semester_suggestion,
    })


def semester_update(request, pk):
    semester = get_object_or_404(Semester, pk=pk)
    if request.method == 'POST':
        form = SemesterForm(request.POST, instance=semester)
        if form.is_valid():
            semester = form.save()
            messages.success(request, "Semester berhasil diperbarui!")
            if semester.aktif:
                messages.info(request, f"{semester.nama} sekarang menjadi semester aktif.")
            return redirect('bendahara:semester_list')
        else:
            messages.error(request, "Form tidak valid, cek kembali input!")
    else:
        form = SemesterForm(instance=semester)
    
    return render(request, 'bendahara/semester_form.html', {
        'form': form,
        'mode': 'edit',
        'semester': semester,
        'semester_suggestion': build_semester_suggestion(),
    })


def semester_delete(request, pk):
    semester = get_object_or_404(Semester, pk=pk)
    if request.method == 'POST':
        semester.delete()
        messages.success(request, "Semester berhasil dihapus!")
        return redirect('bendahara:semester_list')
    
    return render(request, 'bendahara/semester_confirm_delete.html', {
        'semester': semester
    })


def semester_toggle(request, pk):
    semester = get_object_or_404(Semester, pk=pk)
    semester.aktif = True
    semester.save()
    messages.success(request, f"{semester.nama} sekarang menjadi semester aktif.")
    
    return redirect('bendahara:semester_list')


# ================= SISWA =================
def siswa_list(request):
    siswa_list = Siswa.objects.all()
    return render(request, 'bendahara/siswa_list.html', {
        'siswa_list': siswa_list
    })


def siswa_create(request):
    if request.method == 'POST':
        form = SiswaForm(request.POST)

        if form.is_valid():
            siswa = form.save()
            auto_tagihan = create_auto_tagihan_for_siswa(siswa)
            message = "Data siswa berhasil disimpan."
            if auto_tagihan['semester']:
                if auto_tagihan['created_count'] > 0:
                    message += f" {auto_tagihan['created_count']} tagihan otomatis berhasil dibuat."
                else:
                    message += " Tagihan otomatis sudah sinkron dengan semester aktif."
            else:
                message += " Belum ada semester aktif, jadi tagihan otomatis belum dibuat."
            messages.success(request, message)
            return redirect('bendahara:siswa_list')
        else:
            messages.error(request, "Form tidak valid, cek kembali input!")

    else:
        form = SiswaForm()

    return render(request, 'bendahara/siswa_form.html', {
        'form': form
    })


# ================= JENIS PEMBAYARAN =================
def jenis_pembayaran_list(request):
    data = JenisPembayaran.objects.all()
    return render(request, 'bendahara/jenis_pembayaran_list.html', {'data': data})

#post
def jenis_pembayaran_create(request):
    form = JenisPembayaranForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('bendahara:jenis_pembayaran_list')

    return render(request, 'bendahara/jenis_pembayaran_form.html', {
        'form': form
    })
#edit
def jenis_pembayaran_update(request, pk):
    obj = get_object_or_404(JenisPembayaran, pk=pk)
    form = JenisPembayaranForm(request.POST or None, instance=obj)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('bendahara:jenis_pembayaran_list')

    return render(request, 'bendahara/jenis_pembayaran_form.html', {
        'form': form
    })
#delete
def jenis_pembayaran_delete(request, pk):
    obj = get_object_or_404(JenisPembayaran, pk=pk)

    if request.method == 'POST':
        obj.delete()
        return redirect('bendahara:jenis_pembayaran_list')

    return render(request, 'bendahara/confirm_delete.html', {
        'obj': obj
    })
#status
def jenis_pembayaran_toggle(request, pk):
    obj = get_object_or_404(JenisPembayaran, pk=pk)
    obj.aktif = not obj.aktif
    obj.save()
    return redirect('bendahara:jenis_pembayaran_list')


# ================= TAGIHAN =================
def tagihan_list(request):
    semester = get_current_semester(request)

    if semester:
        tagihan_queryset = Tagihan.objects.filter(semester=semester)
    else:
        tagihan_queryset = Tagihan.objects.none()
    
    siswa_objects = (
        Siswa.objects.prefetch_related(
            Prefetch(
                'tagihan_set',
                queryset=tagihan_queryset.select_related('jenis', 'semester').prefetch_related('pembayaran_set'),
            )
        )
        .order_by('nama')
    )

    siswa_list = []
    for siswa in siswa_objects:
        tagihan_items = list(siswa.tagihan_set.all())
        total_nominal = sum(tagihan.nominal for tagihan in tagihan_items)
        total_terbayar = sum(tagihan.total_terbayar for tagihan in tagihan_items)
        total_sisa = max(total_nominal - total_terbayar, 0)

        if not tagihan_items:
            status_tagihan = 'Belum Ada Tagihan'
        elif total_terbayar <= 0:
            status_tagihan = 'Belum Bayar'
        elif total_sisa <= 0:
            status_tagihan = 'Lunas'
        else:
            status_tagihan = 'Cicilan'

        siswa_list.append({
            'id': siswa.id,
            'nis': siswa.nis,
            'nama': siswa.nama,
            'kelas': siswa.kelas,
            'total_tagihan': len(tagihan_items),
            'total_nominal': total_nominal,
            'total_terbayar': total_terbayar,
            'total_sisa': total_sisa,
            'status_tagihan': status_tagihan,
        })

    semester_list = Semester.objects.all()
    
    return render(request, 'bendahara/tagihan_list.html', {
        'siswa_list': siswa_list,
        'semester_list': semester_list,
        'semester_aktif': semester
    })


def tagihan_create(request):
    current_semester = get_current_semester(request, source='POST' if request.method == 'POST' else 'GET')
    if request.method == 'POST':
        siswa_id = request.POST.get('siswa')
        if siswa_id:
            url = redirect('bendahara:tagihan_update', pk=siswa_id)
            if current_semester:
                url['Location'] += f'?{semester_query_param(current_semester)}'
            return url
        messages.error(request, "Silakan pilih siswa terlebih dahulu.")

    return render(request, 'bendahara/tagihan_form.html', {
        'siswa_list': Siswa.objects.filter(aktif=True).order_by('nama'),
        'semester_list': Semester.objects.all(),
        'current_semester': current_semester,
        'mode': 'select',
    })


def tagihan_update(request, pk):
    siswa = get_object_or_404(Siswa, pk=pk)
    current_semester = get_current_semester(request, source='POST' if request.method == 'POST' else 'GET')

    existing_tagihan_all = list(
        Tagihan.objects.filter(siswa=siswa).select_related('jenis', 'semester').prefetch_related('pembayaran_set')
    )
    existing_type_ids = [tagihan.jenis_id for tagihan in existing_tagihan_all]
    jenis_list = get_applicable_jenis_queryset_for_siswa(siswa, include_ids=existing_type_ids)

    semester_list = Semester.objects.all()

    rows = []
    errors = []

    existing_tagihan_current = [
        tagihan
        for tagihan in existing_tagihan_all
        if current_semester and tagihan.semester_id == current_semester.id
    ]
    existing_by_key = {
        (tagihan.jenis_id, tagihan.urutan_periode): tagihan
        for tagihan in existing_tagihan_current
    }

    def build_single_row(jenis, submitted=None):
        existing = existing_by_key.get((jenis.id, 0))
        row = {
            'jenis': jenis,
            'tagihan': existing,
            'aktif': existing is not None,
            'is_bulanan': False,
            'nominal': str(existing.nominal if existing else jenis.nominal_default),
            'periode': existing.periode if existing and existing.periode else '',
            'jatuh_tempo': (
                existing.jatuh_tempo.isoformat()
                if existing and existing.jatuh_tempo
                else ''
            ),
            'catatan': existing.catatan if existing and existing.catatan else '',
        }
        if submitted:
            row.update(submitted)

        active_tagihan = row['aktif'] or existing is not None
        total_nominal = existing.nominal if existing else (int(row['nominal']) if str(row['nominal']).isdigit() and row['aktif'] else 0)
        total_terbayar = existing.total_terbayar if existing else 0
        row['total_nominal'] = total_nominal
        row['total_terbayar'] = total_terbayar
        row['total_sisa'] = existing.sisa_tagihan if existing else total_nominal
        row['status_pembayaran'] = build_tagihan_status(total_nominal, total_terbayar, active_tagihan)
        return row

    def build_monthly_row(jenis, submitted=None):
        resolved_month_count = get_semester_month_span(current_semester, jenis.jumlah_bulan_per_semester)
        period_rows = []
        submitted_by_order = {
            item['urutan_periode']: item
            for item in (submitted or {}).get('period_rows', [])
        }

        for period in build_monthly_periods(current_semester, jenis.jumlah_bulan_per_semester):
            existing = existing_by_key.get((jenis.id, period['urutan_periode']))
            default_data = {
                'urutan_periode': period['urutan_periode'],
                'periode_label': period['label'],
                'tagihan': existing,
                'aktif': existing is not None,
                'nominal': str(existing.nominal if existing else jenis.nominal_default),
                'jatuh_tempo': (
                    existing.jatuh_tempo.isoformat()
                    if existing and existing.jatuh_tempo
                    else (
                        period['jatuh_tempo_default'].isoformat()
                        if period['jatuh_tempo_default']
                        else ''
                    )
                ),
                'catatan': existing.catatan if existing and existing.catatan else '',
            }
            if period['urutan_periode'] in submitted_by_order:
                default_data.update(submitted_by_order[period['urutan_periode']])

            nominal_value = 0
            if str(default_data['nominal']).isdigit() and default_data['aktif']:
                nominal_value = int(default_data['nominal'])

            total_terbayar = existing.total_terbayar if existing else 0
            default_data['total_terbayar'] = total_terbayar
            default_data['total_sisa'] = existing.sisa_tagihan if existing else nominal_value
            default_data['status_pembayaran'] = build_tagihan_status(
                nominal_value,
                total_terbayar,
                default_data['aktif'] or existing is not None,
            )
            period_rows.append(default_data)

        active_rows = [row for row in period_rows if row['aktif'] or row['tagihan']]
        total_nominal = sum(int(row['nominal']) for row in active_rows if str(row['nominal']).isdigit())
        total_terbayar = sum(row['total_terbayar'] for row in period_rows)
        total_sisa = sum(row['total_sisa'] for row in period_rows if row['aktif'] or row['tagihan'])

        return {
            'jenis': jenis,
            'aktif': bool(active_rows),
            'is_bulanan': True,
            'period_rows': period_rows,
            'jumlah_bulan': resolved_month_count,
            'periode_aktif': len([row for row in period_rows if row['aktif']]),
            'total_nominal': total_nominal,
            'total_terbayar': total_terbayar,
            'total_sisa': total_sisa,
            'status_pembayaran': build_tagihan_status(total_nominal, total_terbayar, bool(active_rows)),
        }

    if request.method == 'POST':
        semester_id = request.POST.get('semester')

        if semester_id:
            current_semester = get_object_or_404(Semester, pk=semester_id)
            existing_tagihan_current = [
                tagihan for tagihan in existing_tagihan_all
                if tagihan.semester_id == current_semester.id
            ]
            existing_by_key = {
                (tagihan.jenis_id, tagihan.urutan_periode): tagihan
                for tagihan in existing_tagihan_current
            }

        rows = []
        submitted_actions = []

        for jenis in jenis_list:
            aktif = request.POST.get(f'aktif_{jenis.id}') == '1'

            if jenis.is_bulanan:
                period_rows = []
                active_period_count = 0

                for period in build_monthly_periods(current_semester, jenis.jumlah_bulan_per_semester):
                    existing = existing_by_key.get((jenis.id, period['urutan_periode']))
                    period_active = aktif and request.POST.get(
                        f'aktif_bulan_{jenis.id}_{period["urutan_periode"]}'
                    ) == '1'
                    nominal_raw = request.POST.get(
                        f'nominal_bulan_{jenis.id}_{period["urutan_periode"]}',
                        '',
                    ).strip()
                    jatuh_tempo = request.POST.get(
                        f'jatuh_tempo_bulan_{jenis.id}_{period["urutan_periode"]}',
                        '',
                    ).strip()
                    catatan = request.POST.get(
                        f'catatan_bulan_{jenis.id}_{period["urutan_periode"]}',
                        '',
                    ).strip()

                    period_data = {
                        'urutan_periode': period['urutan_periode'],
                        'periode_label': period['label'],
                        'tagihan': existing,
                        'aktif': period_active,
                        'nominal': nominal_raw if nominal_raw else str(jenis.nominal_default),
                        'jatuh_tempo': jatuh_tempo,
                        'catatan': catatan,
                    }
                    period_rows.append(period_data)

                    if not period_active:
                        continue

                    active_period_count += 1

                    if not nominal_raw:
                        errors.append(f"Nominal untuk {jenis.nama} periode {period['label']} wajib diisi.")
                        continue

                    try:
                        nominal_value = int(nominal_raw)
                    except ValueError:
                        errors.append(f"Nominal untuk {jenis.nama} periode {period['label']} harus berupa angka.")
                        continue

                    if nominal_value <= 0:
                        errors.append(f"Nominal untuk {jenis.nama} periode {period['label']} harus lebih dari 0.")
                        continue

                    if existing and nominal_value < existing.total_terbayar:
                        errors.append(
                            f"Nominal {jenis.nama} periode {period['label']} tidak boleh lebih kecil dari total terbayar Rp {existing.total_terbayar:,}."
                        )
                        continue

                    period_data['nominal_value'] = nominal_value
                    submitted_actions.append({
                        'mode': 'monthly',
                        'jenis': jenis,
                        'periode_label': period['label'],
                        'urutan_periode': period['urutan_periode'],
                        'tagihan': existing,
                        'aktif': period_active,
                        'nominal_value': nominal_value,
                        'jatuh_tempo': jatuh_tempo or None,
                        'catatan': catatan or None,
                    })

                if aktif and active_period_count == 0:
                    errors.append(f"Pilih minimal satu bulan untuk {jenis.nama}.")

                rows.append(build_monthly_row(
                    jenis,
                    submitted={'period_rows': period_rows},
                ))
                continue

            existing = existing_by_key.get((jenis.id, 0))
            nominal_raw = request.POST.get(f'nominal_{jenis.id}', '').strip()
            periode = request.POST.get(f'periode_{jenis.id}', '').strip()
            jatuh_tempo = request.POST.get(f'jatuh_tempo_{jenis.id}', '').strip()
            catatan = request.POST.get(f'catatan_{jenis.id}', '').strip()

            row = build_single_row(jenis, {
                'aktif': aktif,
                'nominal': nominal_raw if nominal_raw else str(jenis.nominal_default),
                'periode': periode,
                'jatuh_tempo': jatuh_tempo,
                'catatan': catatan,
            })
            rows.append(row)

            if not aktif:
                continue

            if not nominal_raw:
                errors.append(f"Nominal untuk {jenis.nama} wajib diisi.")
                continue

            try:
                nominal = int(nominal_raw)
            except ValueError:
                errors.append(f"Nominal untuk {jenis.nama} harus berupa angka.")
                continue

            if nominal <= 0:
                errors.append(f"Nominal untuk {jenis.nama} harus lebih dari 0.")
                continue

            if existing and nominal < existing.total_terbayar:
                errors.append(
                    f"Nominal {jenis.nama} tidak boleh lebih kecil dari total terbayar Rp {existing.total_terbayar:,}."
                )
                continue

            submitted_actions.append({
                'mode': 'single',
                'jenis': jenis,
                'tagihan': existing,
                'aktif': True,
                'nominal_value': nominal,
                'periode': periode or '',
                'jatuh_tempo': jatuh_tempo or None,
                'catatan': catatan or None,
            })

        if not errors and current_semester:
            saved_keys = {
                (action['jenis'].id, action.get('urutan_periode', 0))
                for action in submitted_actions
            }
            for existing in existing_tagihan_current:
                key = (existing.jenis_id, existing.urutan_periode)
                if key in saved_keys:
                    continue
                if not can_delete_tagihan(existing):
                    errors.append(
                        f"Tagihan {existing.jenis.nama} {existing.periode or existing.semester.nama} tidak bisa dinonaktifkan karena sudah memiliki riwayat pembayaran."
                    )

        if not errors and current_semester:
            with transaction.atomic():
                for action in submitted_actions:
                    existing = action['tagihan']

                    defaults = {
                        'nominal': action['nominal_value'],
                        'periode': action.get('periode') or action.get('periode_label', ''),
                        'urutan_periode': action.get('urutan_periode', 0),
                        'jatuh_tempo': action['jatuh_tempo'],
                        'catatan': action['catatan'],
                    }

                    if existing:
                        for field, value in defaults.items():
                            setattr(existing, field, value)
                        existing.save()
                    else:
                        Tagihan.objects.create(
                            siswa=siswa,
                            jenis=action['jenis'],
                            semester=current_semester,
                            **defaults,
                        )

                for existing in existing_tagihan_current:
                    key = (existing.jenis_id, existing.urutan_periode)
                    if key in saved_keys:
                        continue
                    existing.delete()

            if errors:
                rows = []
                for jenis in jenis_list:
                    if jenis.is_bulanan:
                        rows.append(build_monthly_row(jenis))
                    else:
                        rows.append(build_single_row(jenis))
            else:
                messages.success(request, f"Tagihan untuk {siswa.nama} berhasil diperbarui.")
                url = redirect('bendahara:tagihan_update', pk=siswa.pk)
                url['Location'] += f'?{semester_query_param(current_semester)}'
                return url
        elif not current_semester:
            messages.error(request, "Belum ada semester aktif. Buat atau aktifkan semester terlebih dahulu.")
    else:
        for jenis in jenis_list:
            if jenis.is_bulanan:
                rows.append(build_monthly_row(jenis))
            else:
                rows.append(build_single_row(jenis))

    for error in errors:
        messages.error(request, error)

    return render(request, 'bendahara/tagihan_form.html', {
        'mode': 'edit',
        'siswa': siswa,
        'rows': rows,
        'semester_list': semester_list,
        'current_semester': current_semester,
    })


def tagihan_download(request, pk):
    siswa = get_object_or_404(Siswa, pk=pk)
    semester = get_current_semester(request)
    tagihan_queryset = Tagihan.objects.filter(siswa=siswa)
    if semester:
        tagihan_queryset = tagihan_queryset.filter(semester=semester)
    tagihan_list = list(
        tagihan_queryset.select_related('jenis', 'semester').order_by(
            'semester__nama',
            'jenis__nama',
            'urutan_periode',
            'pk',
        )
    )
    total_nominal = sum(tagihan.nominal for tagihan in tagihan_list)
    total_terbayar = sum(tagihan.total_terbayar for tagihan in tagihan_list)
    total_sisa = sum(tagihan.sisa_tagihan for tagihan in tagihan_list)
    tagihan_groups = build_tagihan_display_groups(tagihan_list)

    html = render_to_string('bendahara/tagihan_receipt.html', {
        'siswa': siswa,
        'tagihan_list': tagihan_list,
        'tagihan_groups': tagihan_groups,
        'total_nominal': total_nominal,
        'total_terbayar': total_terbayar,
        'total_sisa': total_sisa,
        'semester': semester,
        'tanggal_cetak': localdate(),
    })

    response = HttpResponse(html, content_type='text/html; charset=utf-8')
    response['Content-Disposition'] = (
        f'attachment; filename="tagihan_{siswa.nis}_{siswa.nama.replace(" ", "_")}.html"'
    )
    return response


def pembayaran_download(request, pk):
    pembayaran = get_object_or_404(
        Pembayaran.objects.select_related(
            'transaksi',
            'tagihan__siswa',
            'tagihan__jenis',
            'tagihan__semester',
        ),
        pk=pk,
    )
    if pembayaran.transaksi_id:
        return render_pembayaran_receipt_response(transaksi=pembayaran.transaksi)

    return render_pembayaran_receipt_response(pembayaran=pembayaran)


def transaksi_pembayaran_download(request, pk):
    transaksi = get_object_or_404(
        TransaksiPembayaran.objects.select_related('siswa', 'semester'),
        pk=pk,
    )

    return render_pembayaran_receipt_response(transaksi=transaksi)


def pembayaran_detail_siswa(request, pk):
    siswa = get_object_or_404(Siswa, pk=pk)
    semester = get_current_semester(request)
    tagihan_queryset = Tagihan.objects.filter(siswa=siswa)
    if semester:
        tagihan_queryset = tagihan_queryset.filter(semester=semester)
    tagihan_list = list(
        tagihan_queryset
        .select_related('jenis', 'semester')
        .prefetch_related('pembayaran_set')
        .order_by('semester__nama', 'jenis__nama', 'urutan_periode', 'pk')
    )

    total_nominal = 0
    total_terbayar = 0

    for tagihan in tagihan_list:
        total_nominal += tagihan.nominal
        total_terbayar += tagihan.total_terbayar
    tagihan_groups = build_tagihan_display_groups(tagihan_list)

    context = {
        'siswa': siswa,
        'tagihan_groups': tagihan_groups,
        'total_nominal': total_nominal,
        'total_terbayar': total_terbayar,
        'total_sisa': max(total_nominal - total_terbayar, 0),
        'semester': semester,
        'semester_query': semester_query_param(semester),
    }
    return render(request, 'bendahara/pembayaran_detail_siswa.html', context)


# ================= PEMBAYARAN =================
def pembayaran_list(request):
    semester = get_current_semester(request)

    if semester:
        pembayaran_list = (
            Pembayaran.objects.select_related(
                'transaksi',
                'tagihan__siswa',
                'tagihan__jenis',
                'tagihan__semester',
            )
            .filter(tagihan__semester=semester)
            .order_by('-tanggal_bayar')
        )
    else:
        pembayaran_list = Pembayaran.objects.none()
    
    semester_list = Semester.objects.all()
    
    return render(request, 'bendahara/pembayaran_list.html', {
        'pembayaran_list': pembayaran_list,
        'semester_list': semester_list,
        'semester_aktif': semester
    })


def pembayaran_create(request):
    siswa_id = request.GET.get('siswa') or request.POST.get('siswa') or request.POST.get('siswa_id')
    tagihan_id = request.GET.get('tagihan') or request.POST.get('tagihan_id')
    current_semester = get_current_semester(request, source='POST' if request.method == 'POST' else 'GET')
    semester_id = request.GET.get('semester') or request.POST.get('semester') or request.POST.get('semester_id')
    if not semester_id and current_semester:
        semester_id = current_semester.pk
    redirect_siswa_id = request.GET.get('redirect_siswa') or request.POST.get('redirect_siswa')
    lock_siswa = bool(
        request.GET.get('siswa') or
        request.POST.get('siswa') or
        request.POST.get('siswa_id') or
        redirect_siswa_id
    )

    selected_siswa = Siswa.objects.filter(pk=siswa_id).first() if siswa_id else None
    selected_semester = Semester.objects.filter(pk=semester_id).first() if semester_id else current_semester

    if request.method == 'POST':
        form = PembayaranMultiForm(request.POST, siswa_id=siswa_id, semester_id=semester_id)
        if lock_siswa:
            form.fields['siswa'].widget = form.fields['siswa'].hidden_widget()

        if form.is_valid():
            selected_siswa = form.cleaned_data['siswa']
            selected_semester = form.cleaned_data['semester'] or current_semester
            available_tagihan = get_available_tagihan_for_payment(selected_siswa, selected_semester)
            pembayaran_groups = build_pembayaran_groups(available_tagihan, data=request.POST)

            errors = []
            selected_rows = []

            if not selected_semester:
                errors.append("Pilih semester terlebih dahulu sebelum menyimpan pembayaran.")

            for group in pembayaran_groups:
                for row in group['rows']:
                    if not row['selected']:
                        continue

                    amount_raw = row['jumlah_input']
                    periode_label = row['tagihan'].periode or row['tagihan'].semester.nama
                    label = (
                        f"{row['tagihan'].jenis.nama} periode {periode_label}"
                        if row['tagihan'].jenis.is_bulanan
                        else row['tagihan'].jenis.nama
                    )

                    if not amount_raw:
                        errors.append(f"Jumlah bayar untuk {label} wajib diisi.")
                        continue

                    try:
                        jumlah_bayar = int(amount_raw)
                    except ValueError:
                        errors.append(f"Jumlah bayar untuk {label} harus berupa angka.")
                        continue

                    if jumlah_bayar <= 0:
                        errors.append(f"Jumlah bayar untuk {label} harus lebih dari 0.")
                        continue

                    if jumlah_bayar > row['tagihan'].sisa_tagihan:
                        errors.append(
                            f"Jumlah bayar untuk {label} melebihi sisa tagihan Rp {row['tagihan'].sisa_tagihan:,}."
                        )
                        continue

                    row['jumlah_bayar'] = jumlah_bayar
                    selected_rows.append(row)

            if not selected_rows:
                errors.append("Pilih minimal satu tagihan yang ingin dibayar.")

            if not errors:
                with transaction.atomic():
                    transaksi = TransaksiPembayaran.objects.create(
                        siswa=selected_siswa,
                        semester=selected_semester,
                        metode=form.cleaned_data['metode'] or None,
                        keterangan=form.cleaned_data['keterangan'] or None,
                    )

                    for row in selected_rows:
                        Pembayaran.objects.create(
                            transaksi=transaksi,
                            tagihan=row['tagihan'],
                            jumlah_bayar=row['jumlah_bayar'],
                            metode=transaksi.metode,
                            keterangan=transaksi.keterangan,
                        )

                messages.success(
                    request,
                    f"Transaksi {transaksi.kode_transaksi} berhasil disimpan untuk {len(selected_rows)} tagihan."
                )
                if redirect_siswa_id:
                    url = redirect('bendahara:pembayaran_detail_siswa', pk=redirect_siswa_id)
                    if selected_semester:
                        url['Location'] += f'?{semester_query_param(selected_semester)}'
                    return url
                url = redirect('bendahara:pembayaran_list')
                if selected_semester:
                    url['Location'] += f'?{semester_query_param(selected_semester)}'
                return url

            for error in errors:
                messages.error(request, error)
        else:
            messages.error(request, "Pembayaran gagal disimpan. Periksa kembali data Anda.")
    else:
        form = PembayaranMultiForm(siswa_id=siswa_id, semester_id=semester_id)
        if lock_siswa:
            form.fields['siswa'].widget = form.fields['siswa'].hidden_widget()

    available_tagihan = []
    if selected_siswa and selected_semester:
        available_tagihan = get_available_tagihan_for_payment(selected_siswa, selected_semester)

    pembayaran_groups = build_pembayaran_groups(
        available_tagihan,
        data=request.POST if request.method == 'POST' else None,
        preselected_tagihan_id=tagihan_id,
    )

    semester_list = Semester.objects.all()
    
    return render(request, 'bendahara/pembayaran_form.html', {
        'form': form,
        'selected_siswa': selected_siswa,
        'selected_semester': selected_semester,
        'pembayaran_groups': pembayaran_groups,
        'lock_siswa': lock_siswa,
        'redirect_siswa_id': redirect_siswa_id,
        'siswa_id': siswa_id,
        'tagihan_id': tagihan_id,
        'semester_id': semester_id,
        'current_semester': selected_semester or current_semester,
        'back_to_detail': redirect_siswa_id is not None,
        'semester_list': semester_list,
    })


def upload_siswa(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']

        try:
            # Try to import pandas, if not available, show error message
            try:
                import pandas as pd
            except ImportError:
                messages.error(request, "Modul pandas tidak tersedia. Silakan instal dengan: pip install pandas")
                return redirect('bendahara:siswa_list')

            df = pd.read_excel(file, dtype=str)
            required_columns = {'nis', 'nama', 'kelas'}
            missing_columns = required_columns - set(df.columns)
            if missing_columns:
                messages.error(
                    request,
                    f"Kolom wajib belum lengkap: {', '.join(sorted(missing_columns))}.",
                )
                return redirect('bendahara:siswa_list')

            imported_count = 0
            updated_count = 0
            empty_row_count = 0
            auto_created_count = 0
            auto_skipped_count = 0

            for _, row in df.iterrows():
                nis = str(row.get('nis', '')).strip()
                nama = str(row.get('nama', '')).strip()
                kelas = str(row.get('kelas', '')).strip()
                pondok = str(row.get('pondok', '')).strip() if 'pondok' in df.columns else ''

                if nis.lower() == 'nan':
                    nis = ''
                if nama.lower() == 'nan':
                    nama = ''
                if kelas.lower() == 'nan':
                    kelas = ''
                if pondok.lower() == 'nan':
                    pondok = ''

                if not nis or not nama or not kelas:
                    empty_row_count += 1
                    continue

                siswa, created = Siswa.objects.update_or_create(
                    nis=nis,
                    defaults={
                        'nama': nama,
                        'kelas': kelas,
                        'pondok': pondok,
                        'aktif': True,
                    },
                )
                if created:
                    imported_count += 1
                else:
                    updated_count += 1

                auto_tagihan = create_auto_tagihan_for_siswa(siswa)
                auto_created_count += auto_tagihan['created_count']
                auto_skipped_count += auto_tagihan['skipped_count']

            message = (
                f"Import selesai. {imported_count} siswa baru ditambahkan, "
                f"{updated_count} siswa diperbarui, "
                f"{auto_created_count} tagihan otomatis dibuat."
            )
            if auto_skipped_count:
                message += f" {auto_skipped_count} tagihan sudah ada."
            if empty_row_count:
                message += f" {empty_row_count} baris dilewati karena data wajib kosong."
            if not get_active_semester():
                message += " Belum ada semester aktif, sehingga tagihan otomatis tidak dibuat."
            messages.success(request, message)

        except Exception as e:
            messages.error(request, f"Terjadi error: {e}")

        return redirect('bendahara:siswa_list')

    return redirect('bendahara:siswa_list')


def siswa_update(request, pk):
    siswa = get_object_or_404(Siswa, pk=pk)

    if request.method == 'POST':
        form = SiswaForm(request.POST, instance=siswa)
        if form.is_valid():
            siswa = form.save()
            auto_tagihan = create_auto_tagihan_for_siswa(siswa)
            message = "Data siswa berhasil diperbarui."
            if auto_tagihan['semester'] and auto_tagihan['created_count'] > 0:
                message += f" {auto_tagihan['created_count']} tagihan baru otomatis ditambahkan."
            elif not auto_tagihan['semester']:
                message += " Belum ada semester aktif, jadi sinkronisasi tagihan otomatis dilewati."
            messages.success(request, message)
            return redirect('bendahara:siswa_list')
    else:
        form = SiswaForm(instance=siswa)

    return render(request, 'bendahara/siswa_form.html', {
        'form': form
    })


def siswa_delete(request, pk):
    siswa = get_object_or_404(Siswa, pk=pk)

    if request.method == 'POST':
        siswa.delete()
        return redirect('bendahara:siswa_list')

    return render(request, 'bendahara/siswa_confirm_delete.html', {
        'siswa': siswa
    })


def download_template_siswa(request):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Template Siswa"

    # Header
    ws.append(["nis", "nama", "kelas", "pondok"])

    # Contoh data
    ws.append(["001", "Budi", "7A", "Pondok A"])
    ws.append(["002", "Siti", "7B", "Pondok B"])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=template_siswa.xlsx'

    wb.save(response)
    return response


# ================= BULK TAGIHAN PER SEMESTER =================
def buat_tagihan_semester(request):
    if request.method == 'POST':
        form = BulkTagihanForm(request.POST)
        if form.is_valid():
            semester = form.cleaned_data['semester']
            jenis_pembayaran_list = form.cleaned_data['jenis_pembayaran']
            nominal = form.cleaned_data['nominal']
            jatuh_tempo = form.cleaned_data['jatuh_tempo']
            
            siswa_aktif = Siswa.objects.filter(aktif=True)
            created_count = 0
            skipped_count = 0
            
            with transaction.atomic():
                for siswa in siswa_aktif:
                    for jenis in jenis_pembayaran_list:
                        if not jenis.applies_to_student(siswa):
                            continue

                        if jenis.is_bulanan:
                            for period in build_monthly_periods(
                                semester,
                                jenis.jumlah_bulan_per_semester,
                                due_date=jatuh_tempo,
                            ):
                                tagihan, created = Tagihan.objects.get_or_create(
                                    siswa=siswa,
                                    jenis=jenis,
                                    semester=semester,
                                    urutan_periode=period['urutan_periode'],
                                    defaults={
                                        'nominal': nominal,
                                        'periode': period['label'],
                                        'jatuh_tempo': period['jatuh_tempo_default'],
                                    }
                                )
                                if created:
                                    created_count += 1
                                else:
                                    skipped_count += 1
                        else:
                            tagihan, created = Tagihan.objects.get_or_create(
                                siswa=siswa,
                                jenis=jenis,
                                semester=semester,
                                urutan_periode=0,
                                defaults={
                                    'nominal': nominal,
                                    'jatuh_tempo': jatuh_tempo,
                                }
                            )
                            if created:
                                created_count += 1
                            else:
                                skipped_count += 1
            
            messages.success(request, f"Berhasil membuat {created_count} tagihan baru. {skipped_count} tagihan sudah ada.")
            url = redirect('bendahara:tagihan_list')
            url['Location'] += f'?{semester_query_param(semester)}'
            return url
        else:
            messages.error(request, "Form tidak valid, cek kembali input!")
    else:
        form = BulkTagihanForm()
    
    return render(request, 'bendahara/buat_tagihan_semester.html', {
        'form': form
    })


# ================= LAPORAN / REKAPAN =================
def build_laporan_semester_context(request):
    semester = get_current_semester(request)

    if semester:
        tagihan_list = Tagihan.objects.filter(semester=semester).select_related('siswa', 'jenis', 'semester')
    else:
        tagihan_list = Tagihan.objects.none()
    
    # Hitung total per jenis pembayaran
    jenis_pembayaran_summary = []
    for jenis in JenisPembayaran.objects.filter(aktif=True):
        tagihan_per_jenis = tagihan_list.filter(jenis=jenis)

        # ✅ WAJIB: ubah ke list supaya bisa pakai property
        tagihan_list_per_jenis = list(tagihan_per_jenis)

        total_nominal = sum(t.nominal for t in tagihan_list_per_jenis)
        total_terbayar = sum(t.total_terbayar for t in tagihan_list_per_jenis)
        total_sisa = total_nominal - total_terbayar

        jumlah_siswa = len(tagihan_list_per_jenis)

        # ✅ FIX: hitung manual (tidak pakai filter)
        jumlah_lunas = sum(1 for t in tagihan_list_per_jenis if t.sisa_tagihan == 0)

        if jumlah_siswa > 0:
            jenis_pembayaran_summary.append({
                'jenis': jenis,
                'total_nominal': total_nominal,
                'total_terbayar': total_terbayar,
                'total_sisa': total_sisa,
                'jumlah_siswa': jumlah_siswa,
                'jumlah_lunas': jumlah_lunas,
                'persen_lunas': (jumlah_lunas / jumlah_siswa * 100)
            })
    
    # Total keseluruhan
    total_nominal = sum(item['total_nominal'] for item in jenis_pembayaran_summary)
    total_terbayar = sum(item['total_terbayar'] for item in jenis_pembayaran_summary)
    total_sisa = sum(item['total_sisa'] for item in jenis_pembayaran_summary)
    
    semester_list = Semester.objects.all()
    
    context = {
        'semester': semester,
        'semester_list': semester_list,
        'jenis_pembayaran_summary': jenis_pembayaran_summary,
        'total_nominal': total_nominal,
        'total_terbayar': total_terbayar,
        'total_sisa': total_sisa,
    }
    
    return context


def laporan_semester(request):
    context = build_laporan_semester_context(request)
    export_response = export_report_response(
        request=request,
        title='Laporan Per Semester',
        filename_prefix='laporan_semester',
        sheet_name='Laporan Semester',
        filter_rows=[
            ('Semester', context['semester'].nama if context['semester'] else 'Belum dipilih'),
        ],
        summary_rows=[
            ('Total Tagihan', format_rupiah(context['total_nominal'])),
            ('Total Terbayar', format_rupiah(context['total_terbayar'])),
            ('Total Tunggakan', format_rupiah(context['total_sisa'])),
        ],
        headers=['Jenis Pembayaran', 'Jumlah Siswa', 'Total Tagihan', 'Total Terbayar', 'Total Sisa', 'Persen Lunas'],
        data_rows=[
            [
                item['jenis'].nama,
                item['jumlah_siswa'],
                format_rupiah(item['total_nominal']),
                format_rupiah(item['total_terbayar']),
                format_rupiah(item['total_sisa']),
                f"{item['persen_lunas']:.1f}%",
            ]
            for item in context['jenis_pembayaran_summary']
        ],
    )
    if export_response:
        return export_response

    return render(request, 'bendahara/laporan_semester.html', context)


def build_laporan_bulanan_context(request):
    today = localdate()
    selected_month = int(request.GET.get('bulan') or today.month)
    selected_year = int(request.GET.get('tahun') or today.year)
    semester = get_current_semester(request)
    kelas_filter = (request.GET.get('kelas') or '').strip()

    payment_queryset = (
        Pembayaran.objects.select_related(
            'transaksi',
            'tagihan__siswa',
            'tagihan__jenis',
            'tagihan__semester',
        )
        .filter(
            tanggal_bayar__year=selected_year,
            tanggal_bayar__month=selected_month,
        )
        .order_by('-tanggal_bayar', '-pk')
    )
    if semester:
        payment_queryset = payment_queryset.filter(tagihan__semester=semester)
    if kelas_filter:
        payment_queryset = payment_queryset.filter(tagihan__siswa__kelas=kelas_filter)

    payment_items = list(payment_queryset)
    payment_report_rows = build_payment_report_rows(payment_items)

    transaction_keys = set()
    siswa_ids = set()
    summary_per_jenis = {}

    for item in payment_report_rows:
        transaction_keys.add(item['kode_transaksi'])
        siswa_ids.add(item['siswa'].id)

        jenis_summary = summary_per_jenis.setdefault(item['jenis'].id, {
            'jenis': item['jenis'],
            'jumlah_item': 0,
            'jumlah_siswa_ids': set(),
            'total_bayar': 0,
        })
        jenis_summary['jumlah_item'] += 1
        jenis_summary['jumlah_siswa_ids'].add(item['siswa'].id)
        jenis_summary['total_bayar'] += item['jumlah_bayar']

    jenis_summary_rows = []
    for item in summary_per_jenis.values():
        jenis_summary_rows.append({
            'jenis': item['jenis'],
            'jumlah_item': item['jumlah_item'],
            'jumlah_siswa': len(item['jumlah_siswa_ids']),
            'total_bayar': item['total_bayar'],
        })
    jenis_summary_rows.sort(key=lambda item: item['total_bayar'], reverse=True)

    year_values = sorted({
        *[dt.year for dt in Pembayaran.objects.dates('tanggal_bayar', 'year')],
        today.year,
    }, reverse=True)
    kelas_list = list(
        Siswa.objects.filter(aktif=True)
        .exclude(kelas__isnull=True)
        .exclude(kelas__exact='')
        .order_by('kelas')
        .values_list('kelas', flat=True)
        .distinct()
    )
    month_options = [
        {'value': index + 1, 'label': name}
        for index, name in enumerate(MONTH_NAMES_ID)
    ]

    context = {
        'semester': semester,
        'semester_list': Semester.objects.all(),
        'selected_month': selected_month,
        'selected_year': selected_year,
        'month_options': month_options,
        'year_options': year_values,
        'kelas_filter': kelas_filter,
        'kelas_list': kelas_list,
        'payment_items': payment_items,
        'payment_report_rows': payment_report_rows,
        'jenis_summary_rows': jenis_summary_rows,
        'total_pembayaran': sum(item.jumlah_bayar for item in payment_items),
        'jumlah_item_pembayaran': len(payment_report_rows),
        'jumlah_transaksi': len(transaction_keys),
        'jumlah_siswa_bayar': len(siswa_ids),
        'bulan_label': MONTH_NAMES_ID[selected_month - 1],
    }
    return context


def laporan_bulanan(request):
    context = build_laporan_bulanan_context(request)
    export_response = export_report_response(
        request=request,
        title='Laporan Per Bulan',
        filename_prefix='laporan_bulanan',
        sheet_name='Laporan Bulanan',
        filter_rows=[
            ('Bulan', f"{context['bulan_label']} {context['selected_year']}"),
            ('Semester', context['semester'].nama if context['semester'] else 'Semua semester'),
            ('Kelas', context['kelas_filter'] or 'Semua kelas'),
        ],
        summary_rows=[
            ('Kas Masuk Bulan Ini', format_rupiah(context['total_pembayaran'])),
            ('Jumlah Transaksi', context['jumlah_transaksi']),
            ('Jumlah Item Pembayaran', context['jumlah_item_pembayaran']),
            ('Jumlah Siswa Membayar', context['jumlah_siswa_bayar']),
            *[
                (f"Kas Masuk {item['jenis'].nama}", format_rupiah(item['total_bayar']))
                for item in context['jenis_summary_rows']
            ],
        ],
        headers=['Tanggal', 'Kode Transaksi', 'NIS', 'Nama Siswa', 'Kelas', 'Jenis', 'Periode', 'Metode', 'Jumlah'],
        data_rows=[
            [
                item['tanggal_bayar'].strftime('%d-%m-%Y %H:%M'),
                item['kode_transaksi'],
                item['siswa'].nis,
                item['siswa'].nama,
                item['siswa'].kelas,
                item['jenis'].nama,
                item['periode_summary'] or '-',
                item['metode'],
                format_rupiah(item['jumlah_bayar']),
            ]
            for item in context['payment_report_rows']
        ],
    )
    if export_response:
        return export_response

    return render(request, 'bendahara/laporan_bulanan.html', context)


def build_laporan_pondok_context(request):
    semester = get_current_semester(request)
    pondok_filter = (request.GET.get('pondok') or '').strip()
    kelas_filter = (request.GET.get('kelas') or '').strip()

    pondok_list = list(
        Siswa.objects.filter(aktif=True)
        .exclude(pondok__exact='')
        .order_by('pondok')
        .values_list('pondok', flat=True)
        .distinct()
    )

    siswa_queryset = Siswa.objects.filter(aktif=True).order_by('pondok', 'kelas', 'nama')
    if pondok_filter:
        siswa_queryset = siswa_queryset.filter(pondok=pondok_filter)
    if kelas_filter:
        siswa_queryset = siswa_queryset.filter(kelas=kelas_filter)

    kelas_list = list(
        siswa_queryset.exclude(kelas__exact='')
        .order_by('kelas')
        .values_list('kelas', flat=True)
        .distinct()
    )

    siswa_list = list(siswa_queryset)
    siswa_ids = [siswa.id for siswa in siswa_list]

    tagihan_queryset = Tagihan.objects.none()
    if semester and siswa_ids:
        tagihan_queryset = (
            Tagihan.objects.filter(siswa_id__in=siswa_ids, semester=semester)
            .select_related('siswa', 'jenis', 'semester')
            .prefetch_related('pembayaran_set')
            .order_by('siswa__pondok', 'siswa__kelas', 'siswa__nama', 'jenis__nama', 'urutan_periode', 'pk')
        )

    tagihan_by_siswa = {}
    for tagihan in tagihan_queryset:
        if tagihan.sisa_tagihan <= 0:
            continue
        tagihan_by_siswa.setdefault(tagihan.siswa_id, []).append(tagihan)

    siswa_pondok_rows = []
    pondok_summary_map = {}
    total_siswa_kurang = 0
    total_item_kurang = 0
    total_sisa = 0

    for siswa in siswa_list:
        outstanding_items = tagihan_by_siswa.get(siswa.id, [])
        if not outstanding_items:
            continue

        siswa_total_sisa = sum(item.sisa_tagihan for item in outstanding_items)
        total_siswa_kurang += 1
        total_item_kurang += len(outstanding_items)
        total_sisa += siswa_total_sisa

        pondok_key = siswa.pondok or 'Belum Diisi'
        if pondok_key not in pondok_summary_map:
            pondok_summary_map[pondok_key] = {
                'pondok': pondok_key,
                'jumlah_siswa_kurang': 0,
                'jumlah_item_kurang': 0,
                'total_sisa': 0,
            }
        pondok_summary_map[pondok_key]['jumlah_siswa_kurang'] += 1
        pondok_summary_map[pondok_key]['jumlah_item_kurang'] += len(outstanding_items)
        pondok_summary_map[pondok_key]['total_sisa'] += siswa_total_sisa

        siswa_pondok_rows.append({
            'siswa': siswa,
            'outstanding_items': outstanding_items,
            'jumlah_item_kurang': len(outstanding_items),
            'detail_kekurangan': format_outstanding_group_details(outstanding_items),
            'total_sisa': siswa_total_sisa,
        })

    pondok_summary_rows = sorted(
        pondok_summary_map.values(),
        key=lambda item: item['total_sisa'],
        reverse=True,
    )

    return {
        'semester': semester,
        'semester_list': Semester.objects.all(),
        'pondok_list': pondok_list,
        'pondok_filter': pondok_filter,
        'kelas_list': kelas_list,
        'kelas_filter': kelas_filter,
        'pondok_summary_rows': pondok_summary_rows,
        'siswa_pondok_rows': siswa_pondok_rows,
        'total_siswa_kurang': total_siswa_kurang,
        'total_item_kurang': total_item_kurang,
        'total_sisa': total_sisa,
    }


def laporan_pondok(request):
    context = build_laporan_pondok_context(request)
    export_response = export_report_response(
        request=request,
        title='Laporan Pondok',
        filename_prefix='laporan_pondok',
        sheet_name='Laporan Pondok',
        filter_rows=[
            ('Semester', context['semester'].nama if context['semester'] else 'Belum dipilih'),
            ('Pondok', context['pondok_filter'] or 'Semua pondok'),
            ('Kelas', context['kelas_filter'] or 'Semua kelas'),
        ],
        summary_rows=[
            ('Jumlah Siswa Kurang Bayar', context['total_siswa_kurang']),
            ('Jumlah Item Kurang', context['total_item_kurang']),
            ('Total Sisa Tagihan', format_rupiah(context['total_sisa'])),
        ],
        headers=['Pondok', 'NIS', 'Nama Siswa', 'Kelas', 'Jumlah Item Kurang', 'Total Sisa', 'Rincian Kekurangan'],
        data_rows=[
            [
                row['siswa'].pondok or '-',
                row['siswa'].nis,
                row['siswa'].nama,
                row['siswa'].kelas,
                row['jumlah_item_kurang'],
                format_rupiah(row['total_sisa']),
                row['detail_kekurangan'],
            ]
            for row in context['siswa_pondok_rows']
        ],
    )
    if export_response:
        return export_response

    return render(request, 'bendahara/laporan_pondok.html', context)


def build_laporan_jenis_pembayaran_context(request):
    semester = get_current_semester(request)
    jenis_queryset = JenisPembayaran.objects.filter(aktif=True).order_by('nama')
    semester_list = Semester.objects.all()
    kelas_list = list(
        Siswa.objects.filter(aktif=True)
        .exclude(kelas__isnull=True)
        .exclude(kelas__exact='')
        .order_by('kelas')
        .values_list('kelas', flat=True)
        .distinct()
    )

    jenis_id = request.GET.get('jenis')
    kelas_filter = request.GET.get('kelas', '').strip()

    selected_jenis = None
    if jenis_id:
        selected_jenis = jenis_queryset.filter(pk=jenis_id).first()
    if selected_jenis is None:
        selected_jenis = jenis_queryset.first()

    siswa_queryset = Siswa.objects.filter(aktif=True).order_by('kelas', 'nama')
    if kelas_filter:
        siswa_queryset = siswa_queryset.filter(kelas=kelas_filter)

    siswa_list = list(siswa_queryset)
    siswa_ids = [siswa.id for siswa in siswa_list]

    tagihan_queryset = Tagihan.objects.none()
    if selected_jenis and semester and siswa_ids:
        tagihan_queryset = (
            Tagihan.objects.filter(
                siswa_id__in=siswa_ids,
                semester=semester,
                jenis=selected_jenis,
            )
            .select_related('siswa', 'jenis', 'semester')
            .prefetch_related('pembayaran_set')
            .order_by('siswa__kelas', 'siswa__nama', 'urutan_periode', 'pk')
        )

    tagihan_per_siswa = {}
    for tagihan in tagihan_queryset:
        tagihan_per_siswa.setdefault(tagihan.siswa_id, []).append(tagihan)

    siswa_status_rows = []
    total_nominal = 0
    total_terbayar = 0
    total_sisa = 0
    jumlah_lunas = 0
    jumlah_belum_lunas = 0
    jumlah_belum_ada_tagihan = 0

    for siswa in siswa_list:
        tagihan_items = tagihan_per_siswa.get(siswa.id, [])
        siswa_total_nominal = sum(tagihan.nominal for tagihan in tagihan_items)
        siswa_total_terbayar = sum(tagihan.total_terbayar for tagihan in tagihan_items)
        siswa_total_sisa = sum(tagihan.sisa_tagihan for tagihan in tagihan_items)
        jumlah_item = len(tagihan_items)
        jumlah_item_lunas = sum(1 for tagihan in tagihan_items if tagihan.sisa_tagihan <= 0)

        if not tagihan_items:
            status = 'Belum Ada Tagihan'
            status_tone = 'slate'
            detail_status = 'Tagihan untuk jenis pembayaran ini belum dibuat.'
            jumlah_belum_ada_tagihan += 1
        elif siswa_total_sisa <= 0:
            status = 'Lunas'
            status_tone = 'green'
            detail_status = (
                f"{jumlah_item_lunas}/{jumlah_item} item lunas."
                if selected_jenis and selected_jenis.is_bulanan
                else 'Pembayaran sudah lunas.'
            )
            jumlah_lunas += 1
        else:
            status = 'Belum Lunas'
            status_tone = 'red' if siswa_total_terbayar <= 0 else 'amber'
            if selected_jenis and selected_jenis.is_bulanan:
                detail_status = f"{jumlah_item_lunas}/{jumlah_item} bulan sudah lunas."
            elif siswa_total_terbayar > 0:
                detail_status = 'Sudah ada pembayaran, tetapi masih ada sisa.'
            else:
                detail_status = 'Belum ada pembayaran masuk.'
            jumlah_belum_lunas += 1

        total_nominal += siswa_total_nominal
        total_terbayar += siswa_total_terbayar
        total_sisa += siswa_total_sisa

        siswa_status_rows.append({
            'siswa': siswa,
            'tagihan_items': tagihan_items,
            'jumlah_item': jumlah_item,
            'jumlah_item_lunas': jumlah_item_lunas,
            'total_nominal': siswa_total_nominal,
            'total_terbayar': siswa_total_terbayar,
            'total_sisa': siswa_total_sisa,
            'status': status,
            'status_tone': status_tone,
            'detail_status': detail_status,
        })

    context = {
        'semester': semester,
        'semester_list': semester_list,
        'jenis_list': jenis_queryset,
        'selected_jenis': selected_jenis,
        'kelas_list': kelas_list,
        'kelas_filter': kelas_filter,
        'siswa_status_rows': siswa_status_rows,
        'total_nominal': total_nominal,
        'total_terbayar': total_terbayar,
        'total_sisa': total_sisa,
        'jumlah_lunas': jumlah_lunas,
        'jumlah_belum_lunas': jumlah_belum_lunas,
        'jumlah_belum_ada_tagihan': jumlah_belum_ada_tagihan,
        'jumlah_siswa': len(siswa_status_rows),
    }
    return context


def laporan_jenis_pembayaran(request):
    context = build_laporan_jenis_pembayaran_context(request)
    export_response = export_report_response(
        request=request,
        title='Laporan Jenis Pembayaran',
        filename_prefix='laporan_jenis_pembayaran',
        sheet_name='Laporan Jenis',
        filter_rows=[
            ('Semester', context['semester'].nama if context['semester'] else 'Belum dipilih'),
            ('Jenis Pembayaran', context['selected_jenis'].nama if context['selected_jenis'] else 'Belum dipilih'),
            ('Kelas', context['kelas_filter'] or 'Semua kelas'),
        ],
        summary_rows=[
            ('Kas Masuk', format_rupiah(context['total_terbayar'])),
            ('Total Tagihan', format_rupiah(context['total_nominal'])),
            ('Total Sisa', format_rupiah(context['total_sisa'])),
            ('Jumlah Lunas', context['jumlah_lunas']),
            ('Jumlah Belum Lunas', context['jumlah_belum_lunas']),
            ('Belum Ada Tagihan', context['jumlah_belum_ada_tagihan']),
        ],
        headers=['NIS', 'Nama Siswa', 'Kelas', 'Status', 'Detail', 'Tagihan', 'Terbayar', 'Sisa'],
        data_rows=[
            [
                row['siswa'].nis,
                row['siswa'].nama,
                row['siswa'].kelas,
                row['status'],
                row['detail_status'],
                format_rupiah(row['total_nominal']),
                format_rupiah(row['total_terbayar']),
                format_rupiah(row['total_sisa']),
            ]
            for row in context['siswa_status_rows']
        ],
    )
    if export_response:
        return export_response

    return render(request, 'bendahara/laporan_jenis_pembayaran.html', context)


def build_laporan_siswa_context(request):
    siswa_id = request.GET.get('siswa')
    semester = get_current_semester(request)
    
    query = Siswa.objects.filter(aktif=True).order_by('kelas', 'nama')
    
    if siswa_id:
        query = query.filter(pk=siswa_id)
    
    semester_list = Semester.objects.all()
    
    siswa_data = []
    for siswa in query:
        tagihan_query = Tagihan.objects.filter(siswa=siswa)
        if semester:
            tagihan_query = tagihan_query.filter(semester=semester)
        
        tagihan_list = list(tagihan_query.select_related('jenis', 'semester'))
        total_nominal = sum(t.nominal for t in tagihan_list)
        total_terbayar = sum(t.total_terbayar for t in tagihan_list)
        total_sisa = total_nominal - total_terbayar
        
        siswa_data.append({
            'siswa': siswa,
            'total_nominal': total_nominal,
            'total_terbayar': total_terbayar,
            'total_sisa': total_sisa,
            'jumlah_tagihan': len(tagihan_list),
        })
    
    context = {
        'siswa_data': siswa_data,
        'semester_list': semester_list,
        'semester_aktif': semester,
        'siswa_selected': get_object_or_404(Siswa, pk=siswa_id) if siswa_id else None,
    }
    
    return context


def laporan_siswa(request):
    context = build_laporan_siswa_context(request)
    export_response = export_report_response(
        request=request,
        title='Laporan Per Siswa',
        filename_prefix='laporan_siswa',
        sheet_name='Laporan Siswa',
        filter_rows=[
            ('Semester', context['semester_aktif'].nama if context['semester_aktif'] else 'Belum dipilih'),
            ('Siswa', context['siswa_selected'].nama if context['siswa_selected'] else 'Semua siswa'),
        ],
        summary_rows=[
            ('Jumlah Baris Siswa', len(context['siswa_data'])),
            ('Total Tagihan', format_rupiah(sum(item['total_nominal'] for item in context['siswa_data']))),
            ('Total Terbayar', format_rupiah(sum(item['total_terbayar'] for item in context['siswa_data']))),
            ('Total Sisa', format_rupiah(sum(item['total_sisa'] for item in context['siswa_data']))),
        ],
        headers=['NIS', 'Nama Siswa', 'Kelas', 'Jumlah Tagihan', 'Total Tagihan', 'Terbayar', 'Sisa'],
        data_rows=[
            [
                item['siswa'].nis,
                item['siswa'].nama,
                item['siswa'].kelas,
                item['jumlah_tagihan'],
                format_rupiah(item['total_nominal']),
                format_rupiah(item['total_terbayar']),
                format_rupiah(item['total_sisa']),
            ]
            for item in context['siswa_data']
        ],
    )
    if export_response:
        return export_response

    return render(request, 'bendahara/laporan_siswa.html', context)


def build_laporan_kelas_context(request):
    semester = get_current_semester(request)
    kelas_list = list(
        Siswa.objects.filter(aktif=True)
        .exclude(kelas__isnull=True)
        .exclude(kelas__exact='')
        .order_by('kelas')
        .values_list('kelas', flat=True)
        .distinct()
    )

    selected_kelas = (request.GET.get('kelas') or '').strip()
    if not selected_kelas and kelas_list:
        selected_kelas = kelas_list[0]

    jenis_list = list(get_applicable_jenis_queryset_for_kelas(selected_kelas))
    siswa_queryset = Siswa.objects.filter(aktif=True).order_by('nama')
    if selected_kelas:
        siswa_queryset = siswa_queryset.filter(kelas=selected_kelas)

    siswa_list = list(siswa_queryset)
    siswa_ids = [siswa.id for siswa in siswa_list]

    tagihan_queryset = Tagihan.objects.none()
    if semester and siswa_ids:
        tagihan_queryset = (
            Tagihan.objects.filter(siswa_id__in=siswa_ids, semester=semester)
            .select_related('siswa', 'jenis', 'semester')
            .prefetch_related('pembayaran_set')
            .order_by('siswa__nama', 'jenis__nama', 'urutan_periode', 'pk')
        )

    tagihan_by_siswa = {}
    for tagihan in tagihan_queryset:
        tagihan_by_siswa.setdefault(tagihan.siswa_id, []).append(tagihan)

    siswa_status_rows = []
    total_nominal = 0
    total_terbayar = 0
    total_sisa = 0

    for siswa in siswa_list:
        tagihan_items = tagihan_by_siswa.get(siswa.id, [])
        summary_groups = build_tagihan_group_summaries(tagihan_items)
        summary_by_jenis_id = {group['jenis'].id: group for group in summary_groups}
        status_columns = []

        for jenis in jenis_list:
            group = summary_by_jenis_id.get(jenis.id)
            if group is None:
                status_columns.append({
                    'jenis': jenis,
                    'status': 'Belum Ada',
                    'status_tone': 'slate',
                    'detail_status': 'Tagihan belum dibuat',
                    'total_sisa': 0,
                })
                continue

            status_columns.append({
                'jenis': jenis,
                'status': group['status'],
                'status_tone': group['status_tone'],
                'detail_status': group['detail_status'],
                'total_sisa': group['total_sisa'],
            })

        siswa_total_nominal = sum(item.nominal for item in tagihan_items)
        siswa_total_terbayar = sum(item.total_terbayar for item in tagihan_items)
        siswa_total_sisa = sum(item.sisa_tagihan for item in tagihan_items)
        total_nominal += siswa_total_nominal
        total_terbayar += siswa_total_terbayar
        total_sisa += siswa_total_sisa

        siswa_status_rows.append({
            'siswa': siswa,
            'tagihan_items': tagihan_items,
            'status_columns': status_columns,
            'rincian_status': format_tagihan_group_details(tagihan_items) if tagihan_items else 'Belum ada tagihan',
            'jumlah_tagihan': len(tagihan_items),
            'total_nominal': siswa_total_nominal,
            'total_terbayar': siswa_total_terbayar,
            'total_sisa': siswa_total_sisa,
        })

    return {
        'semester': semester,
        'semester_list': Semester.objects.all(),
        'kelas_list': kelas_list,
        'selected_kelas': selected_kelas,
        'jenis_list': jenis_list,
        'siswa_status_rows': siswa_status_rows,
        'jumlah_siswa': len(siswa_status_rows),
        'total_nominal': total_nominal,
        'total_terbayar': total_terbayar,
        'total_sisa': total_sisa,
    }


def laporan_kelas(request):
    context = build_laporan_kelas_context(request)
    dynamic_headers = ['NIS', 'Nama Siswa'] + [jenis.nama for jenis in context['jenis_list']] + [
        'Jumlah Tagihan',
        'Total Tagihan',
        'Terbayar',
        'Sisa',
        'Rincian',
    ]
    export_response = export_report_response(
        request=request,
        title='Laporan Per Kelas',
        filename_prefix='laporan_kelas',
        sheet_name='Laporan Kelas',
        filter_rows=[
            ('Semester', context['semester'].nama if context['semester'] else 'Belum dipilih'),
            ('Kelas', context['selected_kelas'] or 'Belum dipilih'),
        ],
        summary_rows=[
            ('Jumlah Siswa', context['jumlah_siswa']),
            ('Total Tagihan', format_rupiah(context['total_nominal'])),
            ('Total Terbayar', format_rupiah(context['total_terbayar'])),
            ('Total Sisa', format_rupiah(context['total_sisa'])),
        ],
        headers=dynamic_headers,
        data_rows=[
            [
                row['siswa'].nis,
                row['siswa'].nama,
                *[column['status'] for column in row['status_columns']],
                row['jumlah_tagihan'],
                format_rupiah(row['total_nominal']),
                format_rupiah(row['total_terbayar']),
                format_rupiah(row['total_sisa']),
                row['rincian_status'],
            ]
            for row in context['siswa_status_rows']
        ],
    )
    if export_response:
        return export_response

    return render(request, 'bendahara/laporan_kelas.html', context)


def build_laporan_tunggakan_context(request):
    semester = get_current_semester(request)

    tagihan_query = Tagihan.objects.all()
    if semester:
        tagihan_query = tagihan_query.filter(semester=semester)
    else:
        tagihan_query = Tagihan.objects.none()
    
    # Group by siswa
    siswa_tunggakan = {}
    for tagihan in tagihan_query.select_related('siswa', 'jenis', 'semester'):
        sisa = tagihan.sisa_tagihan
        if sisa > 0:
            siswa = tagihan.siswa
            if siswa.id not in siswa_tunggakan:
                siswa_tunggakan[siswa.id] = {
                    'siswa': siswa,
                    'total_sisa': 0,
                    'tagihan_items': []
                }
            siswa_tunggakan[siswa.id]['total_sisa'] += sisa
            siswa_tunggakan[siswa.id]['tagihan_items'].append({
                'tagihan': tagihan,
                'sisa': sisa
            })
    
    # Sort by total sisa descending
    tunggakan_list = sorted(siswa_tunggakan.values(), key=lambda x: x['total_sisa'], reverse=True)
    
    semester_list = Semester.objects.all()
    
    context = {
        'tunggakan_list': tunggakan_list,
        'semester_list': semester_list,
        'semester_aktif': semester,
    }
    
    return context


def laporan_tunggakan(request):
    context = build_laporan_tunggakan_context(request)
    export_response = export_report_response(
        request=request,
        title='Laporan Tunggakan',
        filename_prefix='laporan_tunggakan',
        sheet_name='Tunggakan',
        filter_rows=[
            ('Semester', context['semester_aktif'].nama if context['semester_aktif'] else 'Belum dipilih'),
        ],
        summary_rows=[
            ('Jumlah Siswa Menunggak', len(context['tunggakan_list'])),
            ('Total Tunggakan', format_rupiah(sum(item['total_sisa'] for item in context['tunggakan_list']))),
        ],
        headers=['NIS', 'Nama Siswa', 'Kelas', 'Jumlah Tagihan', 'Total Tunggakan', 'Rincian'],
        data_rows=[
            [
                item['siswa'].nis,
                item['siswa'].nama,
                item['siswa'].kelas,
                len(item['tagihan_items']),
                format_rupiah(item['total_sisa']),
                format_outstanding_group_details([detail['tagihan'] for detail in item['tagihan_items']]),
            ]
            for item in context['tunggakan_list']
        ],
    )
    if export_response:
        return export_response

    return render(request, 'bendahara/laporan_tunggakan.html', context)
